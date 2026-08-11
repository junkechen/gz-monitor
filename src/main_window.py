# -*- coding: utf-8 -*-
"""
GZ安环监测系统 - 主界面
使用PyQt5实现
"""

import sys
import os
import time
import gc
import logging
from datetime import datetime, timedelta
from typing import Optional

# ── 日志系统配置 ──────────────────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    # PyInstaller 打包后：日志写到 EXE 同级 logs/ 目录
    _log_dir = os.path.join(os.path.dirname(sys.executable), "logs")
else:
    _log_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "logs")
os.makedirs(_log_dir, exist_ok=True)
_log_file = os.path.join(_log_dir, f"gz_monitor_{datetime.now().strftime('%Y%m%d')}.log")

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(_log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def log_performance(func):
    """性能日志装饰器：记录函数执行时间"""
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        elapsed = time.time() - start
        if elapsed > 1.0:  # 超过1秒才记录
            logger.info(f"[PERF] {func.__name__} 耗时 {elapsed:.2f}秒")
        return result
    return wrapper

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QLabel, QLineEdit, QPushButton, QComboBox, QTableWidget,
    QTableWidgetItem, QListWidget, QListWidgetItem, QHeaderView, QGroupBox, QFormLayout, QMessageBox,
    QDateEdit, QTimeEdit, QFileDialog, QSplitter, QFrame, QScrollArea,
    QDialog, QDialogButtonBox, QSpinBox, QCheckBox, QProgressBar,
    QMenuBar, QMenu, QSizePolicy
)
from PyQt5.QtCore import Qt, QTimer, QDate, QTime, pyqtSignal, QThread
from PyQt5.QtGui import QColor, QFont, QIcon, QPixmap, QPainter, QImage

from multi_account_api import GZMultiAccountClient, GZApiClient
from account_manager import load_accounts, save_accounts, load_warn_thresholds, save_warn_thresholds
from data_processor import (
    parse_realtime_data, parse_history_data, predict_hour_value,
    predict_next_hour_value, predict_day_average, predict_future_hours,
    get_display_params_for_sub, filter_params_for_display,
    is_water_sub, is_gas_sub, get_subtype_code, CODE_TO_NAME,
    get_intervention_state_machine, merge_corrected_data_into_grouped
)
from warning_system import WarningSystem, PredictionWarning
from chart_widget import ChartWidget
from config import (COLORS, WATER_DISPLAY_PARAMS, GAS_DISPLAY_PARAMS, RIGHT_AXIS_PARAMS,
                    HISTORY_TTL, HISTORY_INDEX, SERIES_CAP, FIG_DPI, THROTTLE_MS, EMPTY_PLACEHOLDER)
from refresh_utils import TableDiff, classify_axis
from history_fetch_worker import HistoryFetchWorker, make_history_cache_key
import traceback


# ── 后台预测线程 ──────────────────────────────────────────────────────────────
class PredictionWorker(QThread):
    """
    在后台线程中执行预测计算（API调用 + 预测算法），完成时通过信号传递结果，
    避免在主线程执行导致UI冻结。
    """
    prediction_done = pyqtSignal(list)  # 预测完成后发送全量结果列表

    def __init__(self, grouped_data, multi_client, prediction_horizon,
                 prediction_params, pred_type, thresholds,
                 sub_itemcodes_cache, api_data_cache, thresholds_cache,
                 get_subtype_code_fn, is_water_sub_fn,
                 code_to_name_map, intervention_sm=None,
                 get_sub_itemcodes_fn=None):
        super().__init__()
        self.grouped_data = grouped_data
        self.multi_client = multi_client
        self.prediction_horizon = prediction_horizon
        self.prediction_params = prediction_params
        self.pred_type = pred_type
        self.thresholds = thresholds
        self._sub_itemcodes_cache = sub_itemcodes_cache
        self._api_data_cache = api_data_cache
        self._thresholds_cache = thresholds_cache
        self._get_subtype_code = get_subtype_code_fn
        self._is_water_sub = is_water_sub_fn
        self._code_to_name_map = code_to_name_map
        self._intervention_sm = intervention_sm
        self._get_sub_itemcodes = get_sub_itemcodes_fn

    def run(self):
        """后台线程执行：遍历所有排放口，执行预测，返回全量预测结果"""
        import time as _time
        import traceback

        def _get_cached(sub_key, ttl, fetch_fn, *args, **kwargs):
            now = _time.time()
            entry = self._api_data_cache.get(sub_key)
            if entry is not None:
                ts, data = entry
                if now - ts < ttl:
                    return data
            data = fetch_fn(*args, **kwargs)
            self._api_data_cache[sub_key] = (now, data)
            return data

        all_predictions = []
        total_subs = len(self.grouped_data)
        logger.debug(f"[PRED_WORKER] 开始预测，共 {total_subs} 个排放口，类型: {self.pred_type}")
        
        for unique_key, sub_data in self.grouped_data.items():
            subid = sub_data.get('subid', unique_key)
            sub_type = sub_data.get('subtype', '')
            subtype_code = self._get_subtype_code(sub_type)
            subname = sub_data.get('subname', '')
            ent_name = sub_data.get('ent_name', '')

            # 尝试两种键格式获取监测项目代码
            cache_key_tuple = (subid, subname, subtype_code)
            cache_key_str = f"{subid}_{subtype_code}"
            codes = self._sub_itemcodes_cache.get(cache_key_tuple) or self._sub_itemcodes_cache.get(cache_key_str)
            
            # 如果缓存中没有，尝试主动获取
            if not codes and self._get_sub_itemcodes:
                try:
                    logger.debug(f"[PRED_WORKER] {subname}: 缓存中无监测项目代码，尝试获取...")
                    codes = self._get_sub_itemcodes(subid, subname, subtype_code)
                    if codes:
                        # 更新缓存
                        self._sub_itemcodes_cache[cache_key_str] = codes
                        logger.debug(f"[PRED_WORKER] {subname}: 成功获取监测项目代码: {codes}")
                    else:
                        logger.debug(f"[PRED_WORKER] {subname}: 获取监测项目代码返回空")
                except Exception as e:
                    logger.warning(f"[PRED_WORKER] {subname}: 获取监测项目代码失败: {e}")
                    import traceback
                    logger.warning(f"[PRED_WORKER] 详细错误: {traceback.format_exc()}")
            
            if not codes:
                logger.debug(f"[PRED_WORKER] {subname}: 无监测项目代码，尝试从grouped_data提取...")
                # 备用方案：从grouped_data中直接提取code
                codes_set = set()
                for key, val in group_data.items():
                    if isinstance(val, dict) and 'params' in val:
                        for p in val['params']:
                            c = p.get('code')
                            if c is not None:
                                try:
                                    codes_set.add(str(int(float(c))))
                                except (ValueError, TypeError):
                                    codes_set.add(str(c))
                if codes_set:
                    codes = ','.join(sorted(codes_set, key=lambda x: int(x) if x.isdigit() else 0))
                    logger.debug(f"[PRED_WORKER] {subname}: 从grouped_data提取codes: {codes}")
                else:
                    logger.debug(f"[PRED_WORKER] {subname}: 无监测项目代码，跳过")
                    logger.debug(f"[PRED_WORKER] 可用键: {list(self._sub_itemcodes_cache.keys())[:5]}...")
                    continue

            code_list = [c.strip() for c in codes.split(',') if c.strip()]
            code_to_name = {code: self._code_to_name_map.get(code, code) for code in code_list}
            logger.debug(f"[PRED_WORKER] {subname}: 监测项目 {code_list}")

            client_info = self.multi_client.get_client_by_subid(subid)
            if not client_info:
                logger.debug(f"[PRED_WORKER] {subname}: 无客户端信息，跳过")
                continue
            client = client_info["client"]

            try:
                is_water = self._is_water_sub(sub_type)
                # 判断是否为热电厂——热电厂使用折算数据(showUpload=1)，其他企业使用实测数据
                use_corrected = '热电' in ent_name
                if use_corrected:
                    logger.debug(f"[PRED_WORKER] {ent_name}/{subname}: 热电厂，使用折算数据")
                logger.debug(f"[PRED_WORKER] {subname}: 类型={'废水' if is_water else '废气'}")
                
                if "小时数据预测" in self.pred_type:
                    if is_water:
                        # 废水：只获取一种数据
                        cache_key = f"{subid}_today_hour_{codes}"
                        result = _get_cached(cache_key, 55,
                                             client.get_today_hour_data, subid, subtype_code, codes,
                                             False)
                        rows = result.get('rows', [])
                        corrected_rows = []
                    else:
                        # 废气：热电厂同时获取实测和折算数据，其他只获取实测数据
                        if use_corrected:
                            # 热电厂：同时获取实测和折算
                            cache_key_actual = f"{subid}_minute_{codes}"
                            cache_key_corrected = f"{subid}_minute_{codes}_corrected"
                            result_actual = _get_cached(cache_key_actual, 40,
                                                        client.get_minute_data_current_hour, subid, subtype_code, codes,
                                                        False)
                            result_corrected = _get_cached(cache_key_corrected, 40,
                                                           client.get_minute_data_current_hour, subid, subtype_code, codes,
                                                           True)
                            rows = result_actual.get('rows', [])
                            corrected_rows = result_corrected.get('rows', [])
                        else:
                            # 非热电厂：只获取实测数据
                            cache_key = f"{subid}_minute_{codes}"
                            result = _get_cached(cache_key, 40,
                                                client.get_minute_data_current_hour, subid, subtype_code, codes,
                                                False)
                            rows = result.get('rows', [])
                            corrected_rows = []
                    horizon = self.prediction_horizon

                    for code in code_list:
                        param_name = code_to_name.get(code, code)
                        if self.prediction_params and param_name not in self.prediction_params:
                            logger.debug(f"[PRED_WORKER] {subname}/{param_name}: 不在预测参数列表中，跳过")
                            continue
                        val_key = f"val_{code}"
                        
                        # 对于热电厂废气，使用折算数据预测，同时记录实测数据
                        if use_corrected and corrected_rows:
                            # 热电厂：使用折算数据进行预测
                            # 尝试多种可能的折算字段名
                            def get_corrected_value(row, code):
                                """从行数据中获取折算值，尝试多种字段名"""
                                for key in [f'cvt_{code}', f'Corrected_{code}', f'val_{code}', 'Cvt', 'Corrected', 'cvt']:
                                    val = row.get(key)
                                    if val is not None and val != '':
                                        try:
                                            return float(val)
                                        except (ValueError, TypeError):
                                            pass
                                return None
                            
                            corrected_values = []
                            for r in corrected_rows:
                                cvt_val = get_corrected_value(r, code)
                                if cvt_val is not None and cvt_val != 0:
                                    corrected_values.append(cvt_val)
                            
                            actual_values = []
                            for r in rows:
                                val = r.get(val_key)
                                if val is not None and val != '':
                                    try:
                                        actual_values.append(float(val))
                                    except (ValueError, TypeError):
                                        pass
                            
                            if corrected_values:
                                # IQR 异常值过滤
                                if len(corrected_values) >= 4:
                                    sv = sorted(corrected_values)
                                    q1 = sv[len(sv) // 4]
                                    q3 = sv[3 * len(sv) // 4]
                                    iqr = q3 - q1
                                    lo = q1 - 3.0 * iqr
                                    hi = q3 + 3.0 * iqr
                                    inp = [v for v in corrected_values if lo <= v <= hi]
                                    input_values = inp if inp else corrected_values
                                else:
                                    input_values = corrected_values
                                
                                # 记录实测数据用于显示
                                has_actual_data = len(actual_values) > 0
                                logger.debug(f"[PRED_WORKER] {subname}/{param_name}: 热电厂，使用折算数据预测 ({len(input_values)}点)，实测数据({len(actual_values)}点)")
                            else:
                                logger.debug(f"[PRED_WORKER] {subname}/{param_name}: 热电厂无折算数据，跳过")
                                continue
                        else:
                            # 非热电厂或废水：使用实测数据
                            raw_values = [float(r.get(val_key)) for r in rows
                                          if r.get(val_key) not in (None, '', 0)]
                            logger.debug(f"[PRED_WORKER] {subname}/{param_name}: 获取到 {len(raw_values)} 个有效数据点")
                            if not raw_values:
                                logger.debug(f"[PRED_WORKER] {subname}/{param_name}: 无有效数据，跳过")
                                continue
                            
                            # IQR 异常值过滤
                            if len(raw_values) >= 4:
                                sv = sorted(raw_values)
                                q1 = sv[len(sv) // 4]
                                q3 = sv[3 * len(sv) // 4]
                                iqr = q3 - q1
                                lo = q1 - 3.0 * iqr
                                hi = q3 + 3.0 * iqr
                                inp = [v for v in raw_values if lo <= v <= hi]
                                input_values = inp if inp else raw_values
                            else:
                                input_values = raw_values
                            has_actual_data = False

                        if is_water:
                            from data_processor import predict_day_average, get_warning_level as _gwl2
                            day_res = predict_day_average(input_values, with_trend=True)
                            day_pred = day_res.get('predicted')
                            future_seq = []
                            for offset, label in [(0, '当前小时'), (1, '+1小时'),
                                                   (2, '+2小时'), (3, '+3小时')][:horizon + 1]:
                                future_seq.append({
                                    'hour_offset': offset, 'label': label,
                                    'predicted': day_pred,
                                    'confidence': round(day_res.get('confidence', 0) * (0.85 ** offset), 2),
                                    'trend': day_res.get('trend', 'stable'),
                                    'trend_rate': 0,
                                    'data_points': day_res.get('data_points', 0),
                                    'data_completeness': day_res.get('data_completeness', 0),
                                })
                        else:
                            # ── 干预状态机判断（替代简单的 already_warned）────────────
                            # 根据最近数据趋势自动识别干预状态，而非仅看当前是否超标
                            cur_mean = sum(input_values) / len(input_values) if input_values else 0
                            threshold = self.thresholds.get(param_name, 0) if param_name in self.thresholds else 0

                            # 更新干预状态机，获取干预参数
                            if cur_mean > 0 and threshold > 0:
                                import time
                                intervention_params = self._intervention_sm.update(
                                    subname=subname,
                                    param=param_name,
                                    cur_value=cur_mean,
                                    threshold=threshold,
                                    current_time=time.time()
                                )
                            else:
                                intervention_params = None

                            from data_processor import predict_future_hours as _pfh
                            future_seq = _pfh(input_values, horizon=horizon,
                                            intervention_params=intervention_params)

                        level_order = {"正常": 0, "黄色预警": 1, "橙色预警": 2, "红色预警": 3}
                        max_level = "正常"
                        for fh in future_seq:
                            fval = fh.get('predicted')
                            if fval is None:
                                continue
                            lv = self._get_wl(fval, param_name, self.thresholds)
                            if level_order.get(lv, 0) > level_order.get(max_level, 0):
                                max_level = lv
                        if input_values and param_name in self.thresholds:
                            cur_avg = sum(input_values) / len(input_values)
                            cur_actual = self._get_wl(cur_avg, param_name, self.thresholds)
                            if cur_actual == "正常" and max_level == "红色预警":
                                max_level = "橙色预警"
                                for fh in future_seq:
                                    if fh.get('predicted') is not None:
                                        lv = self._get_wl(fh['predicted'], param_name, self.thresholds)
                                        if lv == "红色预警":
                                            fh['warning_level_cap'] = "橙色预警"

                        cur_item = future_seq[0] if future_seq else {}
                        all_predictions.append({
                            'ent_name': ent_name, 'subname': subname,
                            'param': param_name,
                            'cur_pred': cur_item.get('predicted'),
                            'future_preds': future_seq,
                            'trend': cur_item.get('trend', 'stable'),
                            'trend_rate': cur_item.get('trend_rate', 0),
                            'confidence': cur_item.get('confidence', 0),
                            'data_points': cur_item.get('data_points', 0),
                            'warning_level': max_level,
                            'is_warning': max_level != "正常",
                            'pred_type': '小时预测',
                            'predicted': cur_item.get('predicted'),
                            'use_corrected': use_corrected,  # 标记是否使用折算数据
                            'is_thermal_power': use_corrected and not is_water,  # 标记是否为热电厂废气
                        })
                else:
                    cache_key = f"{subid}_today_hour_{codes}{'_corrected' if use_corrected else ''}"
                    result = _get_cached(cache_key, 55,
                                         client.get_today_hour_data, subid, subtype_code, codes,
                                         use_corrected)
                    rows = result.get('rows', [])
                    from data_processor import predict_day_average as _pda, get_warning_level as _gwl3
                    for code in code_list:
                        param_name = code_to_name.get(code, code)
                        if self.prediction_params and param_name not in self.prediction_params:
                            continue
                        val_key = f"val_{code}"
                        hour_values = [float(r.get(val_key)) for r in rows if r.get(val_key)]
                        if not hour_values:
                            continue
                        pred_res = _pda(hour_values, with_trend=True)
                        predicted = pred_res.get('predicted')
                        wl = _gwl3(predicted, param_name, self.thresholds) if predicted else "正常"
                        all_predictions.append({
                            'ent_name': ent_name, 'subname': subname,
                            'param': param_name, 'cur_pred': predicted,
                            'predicted': predicted,
                            'trend': pred_res.get('trend', 'stable'),
                            'trend_rate': pred_res.get('trend_rate', 0),
                            'confidence': pred_res.get('confidence', 0),
                            'warning_level': wl,
                            'is_warning': wl != "正常",
                            'pred_type': '日均预测',
                        })
            except Exception as e:
                error_detail = traceback.format_exc()
                logger.warning(f"[WARN] 排放口 {subname} 预测失败: {e}")
                logger.warning(f"[WARN] 详细错误: {error_detail}")

        logger.debug(f"[PRED_WORKER] 预测完成，共生成 {len(all_predictions)} 条预测结果")
        self.prediction_done.emit(all_predictions)

    @staticmethod
    def _is_warning_level_not_normal(value, param_name, thresholds):
        from data_processor import get_warning_level as _gwl
        return _gwl(value, param_name, thresholds) != "正常"

    @staticmethod
    def _get_wl(value, param_name, thresholds):
        from data_processor import get_warning_level as _gwl
        return _gwl(value, param_name, thresholds)


# ==================== 版本检测 ====================
def get_app_version():
    """根据程序名判断版本"""
    app_name = os.path.basename(sys.executable if getattr(sys, 'frozen', False) else sys.argv[0]).lower()
    if 'v1' in app_name or 'version1' in app_name:
        return 1
    elif 'v2' in app_name or 'version2' in app_name:
        return 2
    elif 'v3' in app_name or 'version3' in app_name:
        return 3
    return 1  # 默认版本一


# ==================== 登录对话框 ====================
class LoginDialog(QDialog):
    login_success = pyqtSignal(str)  # username

    def __init__(self, parent=None, version=1):
        super().__init__(parent)
        self.version = version
        self.accounts = load_accounts()
        self.current_password = ""
        self.api_client = GZApiClient()
        self.login_progress_label = QLabel("")  # 用于显示多账户登录进度
        self.setWindowTitle("GZ安环 - 一键登录")
        self.setFixedSize(600, 500)
        self.setStyleSheet(f"""
            QDialog {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['bg_dark']}, stop:1 {COLORS['bg_card']});
                color: {COLORS['text_primary']};
            }}
            QGroupBox {{
                border: 2px solid {COLORS['border']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                font-weight: bold;
                color: {COLORS['text_primary']};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: {COLORS['accent']};
            }}
            QTableWidget {{
                background-color: {COLORS['bg_input']};
                border: 1px solid {COLORS['border']};
                border-radius: 4px;
                gridline-color: {COLORS['border']};
            }}
            QTableWidget::item {{
                padding: 8px;
                border: none;
            }}
            QTableWidget::item:selected {{
                background-color: {COLORS['primary']};
                color: white;
            }}
            QHeaderView::section {{
                background-color: {COLORS['bg_dark']};
                color: {COLORS['text_primary']};
                padding: 8px;
                border: 1px solid {COLORS['border']};
                border-bottom: 2px solid {COLORS['primary']};
                font-weight: bold;
            }}
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['primary']}, stop:1 {COLORS['secondary']});
                color: white;
                border: none;
                border-radius: 6px;
                padding: 12px 24px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['secondary']}, stop:1 {COLORS['primary']});
            }}
            QPushButton:pressed {{
                background: {COLORS['primary']};
            }}
            QPushButton:disabled {{
                background: {COLORS['border']};
                color: {COLORS['text_secondary']};
            }}
        """)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)

        # Logo区域 - 增加视觉冲击力
        logo_container = QWidget()
        logo_layout = QVBoxLayout(logo_container)
        logo_layout.setSpacing(5)

        logo_label = QLabel("GZ安环监测系统")
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo_label.setFont(QFont("Microsoft YaHei", 32, QFont.Weight.Bold))
        logo_label.setStyleSheet(f"""
            color: {COLORS['secondary']};
            background: transparent;
            padding: 10px;
        """)
        logo_layout.addWidget(logo_label)

        version_label = QLabel(f"版本 {self.version} | 多企业一键登录")
        version_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        version_label.setStyleSheet(f"""
            color: {COLORS['accent']};
            background: {COLORS['bg_input']};
            padding: 8px 16px;
            border-radius: 15px;
            font-size: 12px;
        """)
        logo_layout.addWidget(version_label)

        layout.addWidget(logo_container)

        # 账户列表展示 - 美化
        account_group = QGroupBox("已配置企业账户（共{}家）".format(len(self.accounts)))
        account_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['primary']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
                color: {COLORS['text_primary']};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: {COLORS['secondary']};
            }}
        """)
        account_layout = QVBoxLayout()

        account_list_widget = QTableWidget()
        account_list_widget.setColumnCount(2)
        account_list_widget.setHorizontalHeaderLabels(["企业名称", "登录状态"])
        account_list_widget.horizontalHeader().setStretchLastSection(True)
        account_list_widget.setRowCount(len(self.accounts))
        account_list_widget.setMaximumHeight(200)
        account_list_widget.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        account_list_widget.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        account_list_widget.horizontalHeader().setStyleSheet(f"""
            QHeaderView::section {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['primary']}, stop:1 {COLORS['secondary']});
                color: white;
                padding: 10px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }}
        """)

        for idx, account in enumerate(self.accounts):
            name_item = QTableWidgetItem(account['name'])
            name_item.setFont(QFont("Microsoft YaHei", 10))
            status_item = QTableWidgetItem("● 待登录")
            status_item.setForeground(QColor(COLORS['text_secondary']))
            status_item.setFont(QFont("Microsoft YaHei", 9))
            account_list_widget.setItem(idx, 0, name_item)
            account_list_widget.setItem(idx, 1, status_item)

        account_layout.addWidget(account_list_widget)
        account_group.setLayout(account_layout)
        layout.addWidget(account_group)

        # 保存引用以便更新状态
        self.account_list_widget = account_list_widget

        # 登录进度显示
        self.status_label = QLabel("点击下方按钮一键登录所有企业账户")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setStyleSheet(f"""
            color: {COLORS['accent']};
            background: {COLORS['bg_input']};
            padding: 12px;
            border-radius: 6px;
            border: 1px solid {COLORS['border']};
        """)
        layout.addWidget(self.status_label)

        # 登录按钮 - 更大更醒目
        self.login_btn = QPushButton("一键登录所有企业")
        self.login_btn.setFont(QFont("Microsoft YaHei", 14, QFont.Weight.Bold))
        self.login_btn.setMinimumHeight(55)
        self.login_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['primary']}, stop:1 {COLORS['secondary']});
                color: white;
                border: 2px solid {COLORS['secondary']};
                border-radius: 10px;
                padding: 15px 30px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['secondary']}, stop:1 {COLORS['primary']});
                border-color: {COLORS['primary']};
            }}
            QPushButton:pressed {{
                background: {COLORS['primary']};
            }}
        """)
        layout.addWidget(self.login_btn)

        # 添加空格
        layout.addStretch()

        # 注：已移除添加/管理账户按钮，仅使用内置企业账户

        self.setLayout(layout)

    def set_login_progress(self, message):
        """设置登录进度消息"""
        self.status_label.setText(f"⏳ {message}")
        self.status_label.setStyleSheet(f"""
            color: {COLORS['accent']};
            background: {COLORS['bg_input']};
            padding: 12px;
            border-radius: 6px;
            border: 1px solid {COLORS['accent']};
        """)
        QApplication.processEvents()

        # 更新账户列表状态
        if self.account_list_widget.rowCount() > 0:
            for row in range(self.account_list_widget.rowCount()):
                status_item = self.account_list_widget.item(row, 1)
                if status_item:
                    status_item.setText("● 登录中...")
                    status_item.setForeground(QColor(COLORS['accent']))

    def set_login_result(self, success, message):
        """设置登录结果"""
        if success:
            self.status_label.setText(f"✅ {message}")
            self.status_label.setStyleSheet(f"""
                color: {COLORS['success']};
                background: {COLORS['bg_input']};
                padding: 12px;
                border-radius: 6px;
                border: 2px solid {COLORS['success']};
                font-weight: bold;
            """)

            # 更新所有账户状态为已登录
            if self.account_list_widget.rowCount() > 0:
                for row in range(self.account_list_widget.rowCount()):
                    status_item = self.account_list_widget.item(row, 1)
                    if status_item:
                        status_item.setText("● 已登录")
                        status_item.setForeground(QColor(COLORS['success']))
        else:
            self.status_label.setText(f"❌ {message}")
            self.status_label.setStyleSheet(f"""
                color: {COLORS['warning']};
                background: {COLORS['bg_input']};
                padding: 12px;
                border-radius: 6px;
                border: 2px solid {COLORS['warning']};
                font-weight: bold;
            """)

    def _do_login(self):
        """执行登录（单账户登录模式，已弃用，保留兼容）"""
        # 新版本使用main.py中的多账户登录流程
        # 这里只保留空实现作为兼容
        pass

    def _change_password(self):
        """修改当前账户密码"""
        username = self.account_combo.currentText()
        if not self._verify_current_password():
            return

        new_pass, ok = self._get_new_password()
        if ok and new_pass:
            for acc in self.accounts:
                if acc['name'] == username:
                    acc['password'] = new_pass
                    break
            save_accounts(self.accounts)
            self.current_password = new_pass
            QMessageBox.information(self, "成功", "密码修改成功！")

    def _verify_current_password(self):
        """验证当前密码"""
        current = self.password_edit.text().strip()
        if current != self.current_password:
            QMessageBox.warning(self, "错误", "当前密码输入错误！")
            return False
        return True

    def _get_new_password(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("修改密码")
        layout = QFormLayout()
        new_edit = QLineEdit()
        new_edit.setEchoMode(QLineEdit.EchoMode.Password)
        confirm_edit = QLineEdit()
        confirm_edit.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow("新密码:", new_edit)
        layout.addRow("确认密码:", confirm_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        dialog.setLayout(layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            new = new_edit.text().strip()
            confirm = confirm_edit.text().strip()
            if not new:
                QMessageBox.warning(self, "错误", "新密码不能为空！")
                return None, False
            if new != confirm:
                QMessageBox.warning(self, "错误", "两次密码不一致！")
                return None, False
            return new, True
        return None, False

    def _add_account(self):
        """添加新账户（版本2）"""
        from config import VERSION2_ADD_PASS, BUILTIN_ACCOUNTS

        if self.version == 2:
            # 检查是否达到上限
            builtin_names = [a['name'] for a in BUILTIN_ACCOUNTS]
            current_extra = [a for a in self.accounts if a['name'] not in builtin_names]
            if len(current_extra) >= 2:
                QMessageBox.warning(self, "提示", "版本二最多可添加2个额外账户！")
                return

            # 验证权限密码
            pass_input, ok = self._get_permission_password("添加账户", VERSION2_ADD_PASS)
            if not ok or not pass_input:
                return

        dialog = QDialog(self)
        dialog.setWindowTitle("添加账户")
        dialog.setFixedSize(400, 200)
        layout = QFormLayout()
        name_edit = QLineEdit()
        pass_edit = QLineEdit()
        pass_edit.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow("企业名称:", name_edit)
        layout.addRow("密码:", pass_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        dialog.setLayout(layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            name = name_edit.text().strip()
            password = pass_edit.text().strip()
            if name and password:
                self.accounts.append({"name": name, "password": password})
                save_accounts(self.accounts)
                self.account_combo.addItem(name)
                QMessageBox.information(self, "成功", "账户添加成功！")

    def _manage_accounts(self):
        """管理账户（版本3）"""
        from config import VERSION3_MANAGE_PASS

        pass_input, ok = self._get_permission_password("账户管理", VERSION3_MANAGE_PASS)
        if not ok or not pass_input:
            return

        # 显示管理对话框
        dialog = AccountManageDialog(self.accounts, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            save_accounts(self.accounts)
            self.account_combo.clear()
            self.account_combo.addItems([a['name'] for a in self.accounts])

    def _get_permission_password(self, title, expected):
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        layout = QVBoxLayout()
        label = QLabel("请输入权限密码:")
        edit = QLineEdit()
        edit.setEchoMode(QLineEdit.EchoMode.Password)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(label)
        layout.addWidget(edit)
        layout.addWidget(buttons)
        dialog.setLayout(layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            return edit.text().strip(), edit.text().strip() == expected
        return None, False


class AccountManageDialog(QDialog):
    """账户管理对话框（版本3）"""

    def __init__(self, accounts: list, parent=None):
        super().__init__(parent)
        self.accounts = accounts
        self.setWindowTitle("账户管理")
        self.setFixedSize(600, 450)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout()
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["企业名称", "密码", "操作"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._load_table()
        layout.addWidget(self.table)

        # 添加按钮
        add_btn = QPushButton("添加账户")
        add_btn.clicked.connect(self._add_account)
        layout.addWidget(add_btn)

        self.setLayout(layout)

    def _load_table(self):
        self.table.setRowCount(len(self.accounts))
        for row, acc in enumerate(self.accounts):
            self.table.setItem(row, 0, QTableWidgetItem(acc['name']))
            self.table.setItem(row, 1, QTableWidgetItem(acc['password']))

            del_btn = QPushButton("删除")
            del_btn.clicked.connect(lambda _, r=row: self._delete_account(r))
            self.table.setCellWidget(row, 2, del_btn)

    def _add_account(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("添加账户")
        dialog.setFixedSize(400, 200)
        layout = QFormLayout()
        name_edit = QLineEdit()
        pass_edit = QLineEdit()
        pass_edit.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow("企业名称:", name_edit)
        layout.addRow("密码:", pass_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        dialog.setLayout(layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            name = name_edit.text().strip()
            password = pass_edit.text().strip()
            if name and password:
                self.accounts.append({"name": name, "password": password})
                self._load_table()

    def _delete_account(self, row):
        reply = QMessageBox.question(self, "确认", "确定删除此账户？",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            del self.accounts[row]
            self._load_table()


# ==================== 主窗口 ====================
class MainWindow(QMainWindow):
    def __init__(self, multi_client: GZMultiAccountClient):
        """
        主窗口
        multi_client: GZMultiAccountClient实例，包含所有登录的企业账户
        """
        try:
            super().__init__()
            self.multi_client = multi_client
        except Exception as e:
            print(f"[ERROR] MainWindow初始化失败: {e}")
            import traceback
            traceback.print_exc()
            raise
        # 收集所有企业名称
        ent_names = list(multi_client.clients.keys())
        self.username = ", ".join(ent_names) if len(ent_names) <= 3 else f"{ent_names[0]}等{len(ent_names)}家企业"
        self.version = get_app_version()
        self.warning_system = WarningSystem()
        self.thresholds = load_warn_thresholds()

        # ── 干预状态机（基于历史趋势自动识别干预状态和强度）─────────────────
        self._intervention_sm = get_intervention_state_machine()
        self.refresh_interval = 1200000  # 默认1200秒（20分钟）刷新

        # 需要预测的参数列表（空列表 = 预测全部参数）
        self.prediction_params = []  # 空列表表示不过滤，预测所有参数

        # ── 缓存与内存管理 ────────────────────────────────────────────────────
        # 预测结果缓存：用于增量更新对比，key=(ent_name, subname, param)，有上限
        self._pred_result_cache: dict = {}
        self._pred_cache_max = 200          # 最多缓存200条预测结果

        # API数据缓存：避免同一个subid重复请求分钟/小时数据，TTL=60秒
        self._api_data_cache: dict = {}     # key=(subid, data_type), value=(timestamp, data)
        self._api_cache_ttl = 60            # 秒

        # 排放口代码缓存（已有）
        self._sub_itemcodes_cache = {}

        # ── 内存管理 ──────────────────────────────────────────────────────────
        self._gc_counter = 0              # GC计数器，每N次刷新执行一次GC
        self._gc_interval = 10             # 每10次刷新执行一次垃圾回收
        self._memory_threshold_mb = 500    # 内存阈值，超过时强制GC（仅提醒）

        logger.info(f"主窗口初始化完成，企业: {self.username}")

        # 预测自动更新相关
        self.auto_update_prediction = True  # 默认开启自动更新预测
        self.prediction_update_interval = 60000  # 预测更新间隔1分钟
        self.prediction_timer = QTimer()
        self.prediction_timer.timeout.connect(self._auto_update_prediction)

        # 声音报警开关（默认开启）
        self.sound_alarm_enabled = True
        # 声音类型（默认 beep1）
        self.sound_alarm_type = 'beep1'

        # 未来小时预测数（小时数据预测时向后预测几个小时，默认3小时）
        self.prediction_horizon = 3

        self._init_ui()
        self._setup_warning()
        self._setup_menu()
        self._refresh_data()  # 初始加载数据

        # 设置自动刷新定时器（三项优化 T01：统一走 _schedule_refresh 入口）
        self.refresh_timer = QTimer()
        self.refresh_timer.timeout.connect(self._schedule_refresh)
        self.refresh_timer.start(self.refresh_interval)

        # ── 三项优化（T01）：历史缓存 / 对比图状态 / 差量器 ──────────────────
        self._history_cache = {}                  # 历史数据缓存（24h TTL），供后台 worker 共享写回
        self._pred_chart_mode = "同排口多参数"     # 对比模式：同排口多参数 / 同参数多排口
        self._pred_chart_normalize = False        # 是否归一化
        self._pred_chart_loading = False          # 预测图加载态
        self._pred_chart_dirty = False            # 预测图按需重绘脏标记
        self._pred_chart_visible = False          # 预测图所在 tab 是否可见
        self._prediction_dirty = False            # 预测触发脏标记（T04 消费）
        self._last_pred_horizon = -1              # 上次预测 horizon（用于预测表列重建判断）
        self._refresh_last_flush = 0.0            # 上次真正刷新时间（节流用，T04）
        self._refresh_timer_coalesce = None       # 节流合并定时器（T04）
        self._hist_worker = None                  # 后台历史取数 worker
        self._hist_results = {}                   # 历史取数结果累加 {subid: {...}}
        self._rt_diff = TableDiff(self.realtime_table)  # 实时表差量渲染器

    def _init_ui(self):
        self.setWindowTitle(f"GZ安环监测系统 - {self.username}")
        self.setMinimumSize(1920, 1080)  # 优化窗口大小,确保显示完整
        self.resize(1920, 1080)  # 设置默认窗口大小

        # 应用主窗口样式
        self.setStyleSheet(f"""
            QMainWindow {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['bg_dark']}, stop:1 {COLORS['bg_card']});
            }}
            QGroupBox {{
                border: 2px solid {COLORS['border']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
                color: {COLORS['text_primary']};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                color: {COLORS['secondary']};
            }}
            QTabWidget::pane {{
                border: 1px solid {COLORS['border']};
                border-radius: 8px;
                background: {COLORS['bg_card']};
            }}
            QTabBar::tab {{
                background: {COLORS['bg_input']};
                color: {COLORS['text_secondary']};
                padding: 12px 24px;
                margin-right: 4px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-weight: bold;
            }}
            QTabBar::tab:selected {{
                background: {COLORS['primary']};
                color: white;
            }}
            QTabBar::tab:hover {{
                background: {COLORS['secondary']};
                color: white;
            }}
            QTableWidget {{
                background-color: {COLORS['bg_input']};
                border: 1px solid {COLORS['border']};
                border-radius: 4px;
                gridline-color: {COLORS['border']};
            }}
            QTableWidget::item {{
                padding: 8px;
            }}
            QTableWidget::item:selected {{
                background-color: {COLORS['primary']};
                color: white;
            }}
            QHeaderView::section {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['primary']}, stop:1 {COLORS['secondary']});
                color: white;
                padding: 10px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }}
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['primary']}, stop:1 {COLORS['secondary']});
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['secondary']}, stop:1 {COLORS['primary']});
            }}
            QLabel {{
                color: {COLORS['text_primary']};
            }}
            QComboBox {{
                background: {COLORS['bg_input']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 4px;
                padding: 6px;
                min-width: 120px;
            }}
            QComboBox::drop-down {{
                border: none;
            }}
            QComboBox::down-arrow {{
                color: {COLORS['accent']};
            }}
            QComboBox QAbstractItemView {{
                background: {COLORS['bg_dark']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                selection-background-color: {COLORS['primary']};
                selection-color: white;
            }}
        """)

        # 主布局
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # ── 左侧容器（收起按钮 + 列表面板）──────────────────────────────────
        left_container = QWidget()
        left_container_layout = QHBoxLayout(left_container)
        left_container_layout.setSpacing(4)
        left_container_layout.setContentsMargins(0, 0, 0, 0)

        # 收起/展开按钮（竖向细条，贴右边）
        self.toggle_btn = QPushButton("◀")
        self.toggle_btn.setFixedWidth(18)
        self.toggle_btn.setSizePolicy(
            self.toggle_btn.sizePolicy().horizontalPolicy(),
            self.toggle_btn.sizePolicy().verticalPolicy()
        )
        self.toggle_btn.setToolTip("收起/展开企业列表")
        self.toggle_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['primary']};
                color: white;
                border: none;
                border-radius: 3px;
                font-size: 9px;
                font-weight: bold;
                padding: 0px;
            }}
            QPushButton:hover {{
                background: {COLORS['secondary']};
            }}
        """)
        self.toggle_btn.clicked.connect(self._toggle_left_panel)

        # 左侧：排放口列表面板
        self.left_panel = QGroupBox("企业排口概览")
        self.left_panel.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['primary']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                color: {COLORS['secondary']};
            }}
        """)
        left_panel_inner = QVBoxLayout()
        self.sub_list = QTableWidget()
        self.sub_list.setColumnCount(2)
        self.sub_list.setHorizontalHeaderLabels(["排放口 (企业名称)", "监测状态"])
        self.sub_list.horizontalHeader().setStretchLastSection(False)  # 取消自动拉伸
        self.sub_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)  # 支持手动调整列宽
        self.sub_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self.sub_list.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.sub_list.itemClicked.connect(self._on_sub_selected)
        # 添加滚动条，确保所有排放口都能显示
        self.sub_list.setMinimumHeight(600)
        # 设置列的初始宽度
        self.sub_list.setColumnWidth(0, 350)  # 排放口列初始宽度
        self.sub_list.setColumnWidth(1, 100)   # 监测状态列初始宽度
        # 确保表头文字完整显示，不自动换行
        self.sub_list.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.sub_list.horizontalHeader().setMinimumSectionSize(80)  # 最小列宽
        self.sub_list.verticalHeader().setVisible(False)
        left_panel_inner.addWidget(self.sub_list)
        self.left_panel.setLayout(left_panel_inner)

        # 设置左侧面板宽度
        self.left_panel.setMinimumWidth(420)
        self.left_panel.setMaximumWidth(420)

        # 组装左侧容器：列表 + 收起按钮
        left_container_layout.addWidget(self.left_panel)
        left_container_layout.addWidget(self.toggle_btn)

        main_layout.addWidget(left_container)

        # 右侧：主内容区
        right_panel = QTabWidget()
        right_tab_style = f"""
            QTabWidget::pane {{
                border: 2px solid {COLORS['primary']};
                border-radius: 10px;
                background: {COLORS['bg_card']};
            }}
        """
        right_panel.setStyleSheet(right_tab_style)

        # Tab 1: 实时监控
        self.realtime_tab = self._create_realtime_tab()
        right_panel.addTab(self.realtime_tab, "📊 实时监控")

        # Tab 2: 历史数据
        self.history_tab = self._create_history_tab()
        right_panel.addTab(self.history_tab, "📈 历史数据")

        # Tab 3: 数据预测
        self.prediction_tab = self._create_prediction_tab()
        right_panel.addTab(self.prediction_tab, "🔮 数据预测")

        # Tab 4: 预警设置
        self.settings_tab = self._create_settings_tab()
        right_panel.addTab(self.settings_tab, "⚙️ 系统设置")

        # 添加tab切换事件处理，确保图表在切换时正确显示
        right_panel.currentChanged.connect(self._on_tab_changed)

        # 右侧添加到主布局
        main_layout.addWidget(right_panel, 1)  # 右侧自动拉伸

        # 状态栏
        self.statusBar().setStyleSheet(f"""
            QStatusBar {{
                background: {COLORS['bg_dark']};
                color: {COLORS['text_primary']};
                border-top: 2px solid {COLORS['primary']};
                padding: 5px;
            }}
        """)
        self._update_status_bar()

    def _create_realtime_tab(self):
        """创建实时监控标签页"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(15, 15, 15, 15)

        # 工具栏
        toolbar_widget = QWidget()
        toolbar = QHBoxLayout(toolbar_widget)
        toolbar.setContentsMargins(0, 0, 0, 0)

        self.last_update_label = QLabel("最后更新: --")
        self.last_update_label.setStyleSheet(f"""
            color: {COLORS['accent']};
            font-size: 14px;
            font-weight: bold;
            padding: 8px 16px;
            background: {COLORS['bg_input']};
            border-radius: 6px;
            border-left: 3px solid {COLORS['accent']};
        """)
        toolbar.addWidget(self.last_update_label)
        toolbar.addStretch()

        refresh_btn = QPushButton("🔄 立即刷新")
        refresh_btn.setMinimumWidth(120)
        refresh_btn.setMinimumHeight(35)
        refresh_btn.clicked.connect(self._refresh_data)
        refresh_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['primary']}, stop:1 {COLORS['secondary']});
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['secondary']}, stop:1 {COLORS['primary']});
            }}
        """)
        toolbar.addWidget(refresh_btn)

        layout.addWidget(toolbar_widget)

        # 排放口信息标题
        self.sub_info_label = QLabel("请选择排放口查看实时数据")
        self.sub_info_label.setStyleSheet(f"""
            color: {COLORS['accent']};
            font-size: 16px;
            font-weight: bold;
            padding: 12px 16px;
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {COLORS['bg_dark']}, stop:1 {COLORS['bg_card']});
            border-radius: 8px;
            border: 2px solid {COLORS['primary']};
        """)
        layout.addWidget(self.sub_info_label)

        # 数据表格
        table_group = QGroupBox("📊 实时监测数据")
        table_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['primary']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                color: {COLORS['secondary']};
            }}
        """)
        table_layout = QVBoxLayout()

        self.realtime_table = QTableWidget()
        self.realtime_table.setColumnCount(7)
        self.realtime_table.setHorizontalHeaderLabels(["监测参数", "实测浓度", "折算浓度", "单位", "排放标准", "折算状态", "监测状态"])
        self.realtime_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.realtime_table.setStyleSheet(f"""
            QTableWidget {{
                background: {COLORS['bg_input']};
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
                gridline-color: {COLORS['border']};
            }}
            QTableWidget::item {{
                padding: 10px;
                font-size: 13px;
            }}
        """)
        table_layout.addWidget(self.realtime_table)
        table_group.setLayout(table_layout)
        layout.addWidget(table_group)

        # 预警状态
        warning_group = QGroupBox("🚨 预警信息")
        warning_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['warning']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                color: {COLORS['warning']};
            }}
        """)
        warning_layout = QVBoxLayout()

        self.warning_list = QLabel("✅ 暂无预警")
        self.warning_list.setStyleSheet(f"""
            color: {COLORS['success']};
            padding: 15px;
            background: {COLORS['bg_dark']};
            border-radius: 6px;
            border: 1px solid {COLORS['border']};
            font-size: 14px;
        """)
        self.warning_list.setWordWrap(True)
        warning_layout.addWidget(self.warning_list)
        warning_group.setLayout(warning_layout)
        layout.addWidget(warning_group)

        widget.setLayout(layout)
        return widget

    def _create_history_tab(self):
        """创建历史数据标签页"""
        widget = QWidget()
        # 使用 stretch=0 固定查询栏，stretch=1 让两个内容区自动填充
        layout = QVBoxLayout()
        layout.setSpacing(8)
        layout.setContentsMargins(15, 15, 15, 15)

        # ── 查询条件（固定高度，不拉伸）───────────────────────────────────
        query_group = QGroupBox("🔍 查询条件")
        query_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['accent']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                color: {COLORS['accent']};
            }}
        """)
        # 使用两行布局：第一行为主要条件，第二行为快捷按钮
        query_v_layout = QVBoxLayout()
        query_v_layout.setSpacing(6)

        # ── 第一行：排放口 + 时间类型 + 时间范围 + 操作按钮 ────────────────
        query_layout = QHBoxLayout()
        query_layout.setSpacing(8)

        query_layout.addWidget(QLabel("排放口:"))
        self.history_sub_combo = QComboBox()
        self.history_sub_combo.setMinimumWidth(220)
        query_layout.addWidget(self.history_sub_combo)

        query_layout.addWidget(QLabel("数据类型:"))
        self.time_type_combo = QComboBox()
        self.time_type_combo.setMinimumWidth(90)
        self.time_type_combo.addItems(["分钟数据", "小时数据", "日数据", "月数据", "季度数据", "年数据"])
        self.time_type_combo.setCurrentIndex(1)  # 默认小时数据
        self.time_type_combo.currentIndexChanged.connect(self._on_time_type_changed)
        query_layout.addWidget(self.time_type_combo)

        # 开始时间
        _dt_style = f"""
            background: {COLORS['bg_input']};
            color: {COLORS['text_primary']};
            border: 1px solid {COLORS['border']};
            border-radius: 4px;
            padding: 4px;
        """
        self._start_lbl = QLabel("开始:")
        query_layout.addWidget(self._start_lbl)
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-7))
        self.start_date.setDisplayFormat("yyyy-MM-dd")
        self.start_date.setCalendarPopup(True)
        self.start_date.setStyleSheet(f"QDateEdit {{ {_dt_style} }}")
        query_layout.addWidget(self.start_date)

        self.start_time = QTimeEdit()
        self.start_time.setTime(QTime(0, 0))
        self.start_time.setDisplayFormat("HH:mm")
        self.start_time.setStyleSheet(f"QTimeEdit {{ {_dt_style} }}")
        query_layout.addWidget(self.start_time)

        # 结束时间
        self._end_lbl = QLabel("结束:")
        query_layout.addWidget(self._end_lbl)
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setDisplayFormat("yyyy-MM-dd")
        self.end_date.setCalendarPopup(True)
        self.end_date.setStyleSheet(f"QDateEdit {{ {_dt_style} }}")
        query_layout.addWidget(self.end_date)

        self.end_time = QTimeEdit()
        self.end_time.setTime(QTime(23, 59))
        self.end_time.setDisplayFormat("HH:mm")
        self.end_time.setStyleSheet(f"QTimeEdit {{ {_dt_style} }}")
        query_layout.addWidget(self.end_time)

        query_btn = QPushButton("🔍 查询")
        query_btn.clicked.connect(self._query_history)
        query_btn.setMinimumWidth(75)
        query_layout.addWidget(query_btn)

        export_btn = QPushButton("📥 导出Excel")
        export_btn.clicked.connect(self._export_history)
        export_btn.setMinimumWidth(95)
        query_layout.addWidget(export_btn)

        query_v_layout.addLayout(query_layout)

        # ── 第二行：快捷时间范围按钮 ─────────────────────────────────────
        quick_layout = QHBoxLayout()
        quick_layout.setSpacing(6)
        quick_layout.addWidget(QLabel("快速选择:"))

        _quick_btn_style = f"""
            QPushButton {{
                background: {COLORS.get('bg_input', '#2c3e50')};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 4px;
                padding: 3px 10px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background: {COLORS['accent']};
                color: white;
            }}
        """
        quick_ranges = [
            ("今天",    "today"),
            ("昨天",    "yesterday"),
            ("近7天",   "7d"),
            ("近30天",  "30d"),
            ("本月",    "this_month"),
            ("上月",    "last_month"),
            ("本季度",  "this_quarter"),
            ("本年",    "this_year"),
            ("上一年",  "last_year"),
        ]
        for label, key in quick_ranges:
            btn = QPushButton(label)
            btn.setStyleSheet(_quick_btn_style)
            btn.setFixedHeight(24)
            btn.clicked.connect(lambda _, k=key: self._quick_set_time_range(k))
            quick_layout.addWidget(btn)

        quick_layout.addStretch()
        query_v_layout.addLayout(quick_layout)

        query_group.setLayout(query_v_layout)
        # 查询栏固定高度，不参与弹性分配
        query_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        layout.addWidget(query_group, 0)

        # ── 表格区域（含折叠标题栏）────────────────────────────────────────
        self.table_section = QWidget()
        self.table_section.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        table_section_layout = QVBoxLayout(self.table_section)
        table_section_layout.setSpacing(4)
        table_section_layout.setContentsMargins(0, 0, 0, 0)

        # 表格折叠标题栏
        table_header_bar = QWidget()
        table_header_bar.setFixedHeight(32)
        table_header_bar.setStyleSheet(f"""
            QWidget {{
                background: {COLORS['primary']};
                border-radius: 6px;
            }}
        """)
        table_header_bar_layout = QHBoxLayout(table_header_bar)
        table_header_bar_layout.setContentsMargins(12, 0, 12, 0)
        table_header_bar_layout.setSpacing(6)

        table_title_lbl = QLabel("📋 历史数据表格")
        table_title_lbl.setStyleSheet("color: white; font-weight: bold; font-size: 13px; background: transparent;")
        self.table_toggle_btn = QPushButton("▲ 收起")
        self.table_toggle_btn.setFixedSize(70, 22)
        self.table_toggle_btn.setStyleSheet(f"""
            QPushButton {{
                background: rgba(255,255,255,0.15);
                color: white;
                border: 1px solid rgba(255,255,255,0.4);
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: rgba(255,255,255,0.3);
            }}
        """)
        self.table_toggle_btn.clicked.connect(self._toggle_history_table)
        table_header_bar_layout.addWidget(table_title_lbl)
        table_header_bar_layout.addStretch()
        table_header_bar_layout.addWidget(self.table_toggle_btn)

        # 表格内容区（可折叠）
        self.table_content = QWidget()
        self.table_content.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        table_content_layout = QVBoxLayout(self.table_content)
        table_content_layout.setContentsMargins(0, 4, 0, 0)
        table_content_layout.setSpacing(0)

        # 数据表格
        table_group = QGroupBox()
        table_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['primary']};
                border-radius: 6px;
                margin-top: 0px;
                padding-top: 6px;
            }}
        """)
        table_inner = QVBoxLayout()
        table_inner.setContentsMargins(6, 6, 6, 6)
        table_inner.setSpacing(4)
        self.history_table = QTableWidget()
        self.history_table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.history_table.setStyleSheet(f"""
            QTableWidget {{
                background: {COLORS['bg_input']};
                border: none;
                border-radius: 4px;
            }}
        """)
        table_inner.addWidget(self.history_table)

        # 分页工具栏
        page_bar = QHBoxLayout()
        page_bar.setContentsMargins(0, 2, 0, 2)

        self.page_prev_btn = QPushButton("◀ 上一页")
        self.page_prev_btn.setFixedWidth(90)
        self.page_prev_btn.setEnabled(False)
        self.page_prev_btn.clicked.connect(self._history_prev_page)
        self.page_prev_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['primary']};
                color: white;
                border-radius: 4px;
                padding: 4px 8px;
            }}
            QPushButton:disabled {{ background: {COLORS['border']}; color: #666; }}
        """)

        self.page_info_label = QLabel("第 0 页 / 共 0 页  (共 0 条)")
        self.page_info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.page_info_label.setStyleSheet(f"color: {COLORS['text_primary']}; font-size: 12px;")

        self.page_next_btn = QPushButton("下一页 ▶")
        self.page_next_btn.setFixedWidth(90)
        self.page_next_btn.setEnabled(False)
        self.page_next_btn.clicked.connect(self._history_next_page)
        self.page_next_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['primary']};
                color: white;
                border-radius: 4px;
                padding: 4px 8px;
            }}
            QPushButton:disabled {{ background: {COLORS['border']}; color: #666; }}
        """)

        page_bar.addStretch()
        page_bar.addWidget(self.page_prev_btn)
        page_bar.addWidget(self.page_info_label)
        page_bar.addWidget(self.page_next_btn)
        page_bar.addStretch()
        table_inner.addLayout(page_bar)

        table_group.setLayout(table_inner)
        table_content_layout.addWidget(table_group)

        table_section_layout.addWidget(table_header_bar)
        table_section_layout.addWidget(self.table_content, 1)  # 内容区弹性撑满
        layout.addWidget(self.table_section, 1)

        # ── 图表区域（含折叠标题栏）────────────────────────────────────────
        self.chart_section = QWidget()
        self.chart_section.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        chart_section_layout = QVBoxLayout(self.chart_section)
        chart_section_layout.setSpacing(4)
        chart_section_layout.setContentsMargins(0, 0, 0, 0)

        # 图表折叠标题栏
        chart_header_bar = QWidget()
        chart_header_bar.setFixedHeight(32)
        chart_header_bar.setStyleSheet(f"""
            QWidget {{
                background: {COLORS['secondary']};
                border-radius: 6px;
            }}
        """)
        chart_header_bar_layout = QHBoxLayout(chart_header_bar)
        chart_header_bar_layout.setContentsMargins(12, 0, 12, 0)
        chart_header_bar_layout.setSpacing(6)

        chart_title_lbl = QLabel("📈 数据趋势图")
        chart_title_lbl.setStyleSheet("color: white; font-weight: bold; font-size: 13px; background: transparent;")
        self.chart_toggle_btn = QPushButton("▲ 收起")
        self.chart_toggle_btn.setFixedSize(70, 22)
        self.chart_toggle_btn.setStyleSheet(f"""
            QPushButton {{
                background: rgba(255,255,255,0.15);
                color: white;
                border: 1px solid rgba(255,255,255,0.4);
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: rgba(255,255,255,0.3);
            }}
        """)
        self.chart_toggle_btn.clicked.connect(self._toggle_history_chart)
        chart_header_bar_layout.addWidget(chart_title_lbl)
        chart_header_bar_layout.addStretch()
        chart_header_bar_layout.addWidget(self.chart_toggle_btn)

        # 图表内容区（可折叠）
        self.chart_content = QWidget()
        self.chart_content.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        chart_content_layout = QVBoxLayout(self.chart_content)
        chart_content_layout.setContentsMargins(0, 4, 0, 0)
        chart_content_layout.setSpacing(0)

        chart_group = QGroupBox()
        chart_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['secondary']};
                border-radius: 6px;
                margin-top: 0px;
                padding-top: 6px;
            }}
        """)
        chart_inner = QVBoxLayout()
        chart_inner.setContentsMargins(6, 6, 6, 6)
        self.chart_widget = ChartWidget()
        self.chart_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        # 去掉 setMinimumHeight，让图表自由伸缩
        chart_inner.addWidget(self.chart_widget)
        chart_group.setLayout(chart_inner)
        chart_content_layout.addWidget(chart_group)

        chart_section_layout.addWidget(chart_header_bar)
        chart_section_layout.addWidget(self.chart_content, 1)  # 内容区弹性撑满
        layout.addWidget(self.chart_section, 1)

        widget.setLayout(layout)
        return widget

    def _create_prediction_tab(self):
        """创建数据预测标签页"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(8)
        layout.setContentsMargins(15, 15, 15, 15)

        # ── 控制栏（固定高度，不拉伸）─────────────────────────────────────
        pred_group = QGroupBox("🔮 预测设置")
        pred_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['accent']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                color: {COLORS['accent']};
            }}
        """)
        pred_layout = QHBoxLayout()
        pred_layout.setSpacing(12)

        # 预测类型
        pred_layout.addWidget(QLabel("预测类型:"))
        self.pred_type_combo = QComboBox()
        self.pred_type_combo.addItems(["小时数据预测", "当日均值预测"])
        self.pred_type_combo.currentIndexChanged.connect(self._on_pred_type_changed)
        pred_layout.addWidget(self.pred_type_combo)

        pred_layout.addSpacing(10)

        # 预测指标选择按钮
        param_btn = QPushButton("📊 预测指标")
        param_btn.setToolTip("选择需要预测的指标（不选则预测全部）")
        param_btn.setMinimumWidth(100)
        param_btn.clicked.connect(self._show_prediction_params_dialog)
        pred_layout.addWidget(param_btn)

        pred_layout.addSpacing(10)

        # 开始预测按钮
        pred_btn = QPushButton("🚀 开始预测")
        pred_btn.clicked.connect(self._run_prediction)
        pred_btn.setMinimumWidth(120)
        pred_layout.addWidget(pred_btn)

        pred_layout.addSpacing(10)

        # 自动更新开关
        chk_style = f"""
            QCheckBox {{
                color: {COLORS['text_primary']};
                font-size: 14px;
                font-weight: bold;
                padding: 8px;
            }}
            QCheckBox::indicator {{
                width: 20px; height: 20px;
            }}
            QCheckBox::indicator:checked {{
                background: {COLORS['primary']};
                border: 2px solid {COLORS['primary']};
            }}
        """
        self.auto_update_checkbox = QCheckBox("🔄 自动更新")
        self.auto_update_checkbox.setChecked(self.auto_update_prediction)
        self.auto_update_checkbox.setStyleSheet(chk_style)
        self.auto_update_checkbox.stateChanged.connect(self._on_auto_update_changed)
        pred_layout.addWidget(self.auto_update_checkbox)

        pred_layout.addSpacing(10)

        # 声音报警开关
        self.sound_alarm_checkbox = QCheckBox("🔔 声音报警")
        self.sound_alarm_checkbox.setChecked(self.sound_alarm_enabled)
        self.sound_alarm_checkbox.setStyleSheet(chk_style)
        self.sound_alarm_checkbox.stateChanged.connect(self._on_sound_alarm_changed)
        pred_layout.addWidget(self.sound_alarm_checkbox)

        # 声音类型下拉框
        self.sound_type_combo = QComboBox()
        self.sound_type_combo.addItems(["叮咚（提示音）", "滴滴（急促音）", "警报（连续音）"])
        self.sound_type_combo.setCurrentIndex(0)
        self.sound_type_combo.setToolTip("选择预警声音类型")
        self.sound_type_combo.setFixedWidth(120)
        self.sound_type_combo.currentIndexChanged.connect(self._on_sound_type_changed)
        pred_layout.addWidget(self.sound_type_combo)

        pred_layout.addSpacing(15)

        # 预测未来小时数（仅小时预测模式有效）
        pred_layout.addWidget(QLabel("预测未来:"))
        self.pred_horizon_combo = QComboBox()
        self.pred_horizon_combo.addItems(["1小时", "2小时", "3小时"])
        self.pred_horizon_combo.setCurrentIndex(self.prediction_horizon - 1)
        self.pred_horizon_combo.currentIndexChanged.connect(self._on_pred_horizon_changed)
        self.pred_horizon_combo.setToolTip("小时数据预测时，向后预测几个小时")
        pred_layout.addWidget(self.pred_horizon_combo)

        pred_layout.addStretch()
        pred_group.setLayout(pred_layout)
        pred_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        layout.addWidget(pred_group, 0)

        # ── 预测结果区域（含折叠标题栏）───────────────────────────────────
        self.pred_result_section = QWidget()
        self.pred_result_section.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        pred_result_layout = QVBoxLayout(self.pred_result_section)
        pred_result_layout.setSpacing(4)
        pred_result_layout.setContentsMargins(0, 0, 0, 0)

        # 折叠标题栏
        pred_result_header = QWidget()
        pred_result_header.setFixedHeight(32)
        pred_result_header.setStyleSheet(f"""
            QWidget {{
                background: {COLORS['primary']};
                border-radius: 6px;
            }}
        """)
        pred_result_header_layout = QHBoxLayout(pred_result_header)
        pred_result_header_layout.setContentsMargins(12, 0, 12, 0)
        pred_result_header_layout.setSpacing(6)

        pred_result_title = QLabel("📊 预测结果（全部排放口）")
        pred_result_title.setStyleSheet("color: white; font-weight: bold; font-size: 13px; background: transparent;")
        self.pred_result_toggle_btn = QPushButton("▲ 收起")
        self.pred_result_toggle_btn.setFixedSize(70, 22)
        self.pred_result_toggle_btn.setStyleSheet(f"""
            QPushButton {{
                background: rgba(255,255,255,0.15);
                color: white;
                border: 1px solid rgba(255,255,255,0.4);
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: rgba(255,255,255,0.3);
            }}
        """)
        self.pred_result_toggle_btn.clicked.connect(self._toggle_pred_result)
        pred_result_header_layout.addWidget(pred_result_title)
        pred_result_header_layout.addStretch()
        pred_result_header_layout.addWidget(self.pred_result_toggle_btn)

        # 内容区（可折叠）
        self.pred_result_content = QWidget()
        self.pred_result_content.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        pred_result_content_layout = QVBoxLayout(self.pred_result_content)
        pred_result_content_layout.setContentsMargins(0, 4, 0, 0)
        pred_result_content_layout.setSpacing(0)

        result_group = QGroupBox()
        result_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['primary']};
                border-radius: 6px;
                margin-top: 0px;
                padding-top: 6px;
            }}
        """)
        result_inner = QVBoxLayout()
        result_inner.setContentsMargins(6, 6, 6, 6)
        result_inner.setSpacing(4)

        # 结果表格（列头由 _rebuild_pred_table_columns 在预测执行前动态生成）
        self.pred_table = QTableWidget()
        self.pred_table.setColumnCount(0)
        self.pred_table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.pred_table.setStyleSheet(f"""
            QTableWidget {{
                background: {COLORS['bg_input']};
                border: none;
                border-radius: 4px;
                gridline-color: {COLORS['border']};
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {COLORS['bg_card']};
                color: {COLORS['text_primary']};
                padding: 6px;
                border: 1px solid {COLORS['border']};
                font-weight: bold;
            }}
            QTableWidget::item {{
                padding: 4px 8px;
                color: {COLORS['text_primary']};
            }}
        """)
        result_inner.addWidget(self.pred_table)

        # 旧的 label 保留（日均预测时展示额外信息）
        self.pred_result_label = QLabel('点击"开始预测"查看预测结果（将对所有排放口的所有参数进行预测）')
        self.pred_result_label.setWordWrap(True)
        self.pred_result_label.setStyleSheet(f"""
            color: {COLORS['text_secondary']};
            padding: 8px 15px;
            font-size: 12px;
        """)
        result_inner.addWidget(self.pred_result_label)

        result_group.setLayout(result_inner)
        pred_result_content_layout.addWidget(result_group)

        pred_result_layout.addWidget(pred_result_header)
        pred_result_layout.addWidget(self.pred_result_content, 1)
        layout.addWidget(self.pred_result_section, 1)

        # ── 预测图表区域（含折叠标题栏）───────────────────────────────────
        self.pred_chart_section = QWidget()
        self.pred_chart_section.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        pred_chart_layout = QVBoxLayout(self.pred_chart_section)
        pred_chart_layout.setSpacing(4)
        pred_chart_layout.setContentsMargins(0, 0, 0, 0)

        # 折叠标题栏
        pred_chart_header = QWidget()
        pred_chart_header.setFixedHeight(32)
        pred_chart_header.setStyleSheet(f"""
            QWidget {{
                background: {COLORS['secondary']};
                border-radius: 6px;
            }}
        """)
        pred_chart_header_layout = QHBoxLayout(pred_chart_header)
        pred_chart_header_layout.setContentsMargins(12, 0, 12, 0)
        pred_chart_header_layout.setSpacing(6)

        pred_chart_title = QLabel("📈 预测趋势图")
        pred_chart_title.setStyleSheet("color: white; font-weight: bold; font-size: 13px; background: transparent;")

        # 三项优化 T02：原单参数下拉框 pred_chart_param_combo 已移除，
        # 改为下方内容区的可勾选"对比参数"面板（多指标叠加，见 pred_compare_panel）。


        self.pred_chart_toggle_btn = QPushButton("▲ 收起")
        self.pred_chart_toggle_btn.setFixedSize(70, 22)
        self.pred_chart_toggle_btn.setStyleSheet(f"""
            QPushButton {{
                background: rgba(255,255,255,0.15);
                color: white;
                border: 1px solid rgba(255,255,255,0.4);
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: rgba(255,255,255,0.3);
            }}
        """)
        self.pred_chart_toggle_btn.clicked.connect(self._toggle_pred_chart)
        pred_chart_header_layout.addWidget(pred_chart_title)
        pred_chart_header_layout.addSpacing(12)
        pred_chart_header_layout.addStretch()
        pred_chart_header_layout.addWidget(self.pred_chart_toggle_btn)

        # 内容区（可折叠）
        self.pred_chart_content = QWidget()
        self.pred_chart_content.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        pred_chart_content_layout = QVBoxLayout(self.pred_chart_content)
        pred_chart_content_layout.setContentsMargins(0, 4, 0, 0)
        pred_chart_content_layout.setSpacing(0)

        # ── 三项优化 T02：多指标对比参数面板（可勾选 QListWidget）────────────
        self.pred_compare_panel = QWidget()
        compare_layout = QHBoxLayout(self.pred_compare_panel)
        compare_layout.setContentsMargins(4, 4, 4, 4)
        compare_layout.setSpacing(8)

        param_lbl = QLabel("对比参数:")
        param_lbl.setStyleSheet(f"color: {COLORS['text_primary']}; font-size: 12px;")
        compare_layout.addWidget(param_lbl)

        self.pred_param_list = QListWidget()
        self.pred_param_list.setFlow(QListWidget.LeftToRight)
        self.pred_param_list.setWrapping(False)
        self.pred_param_list.setSpacing(4)
        self.pred_param_list.setFixedHeight(40)
        self.pred_param_list.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.pred_param_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.pred_param_list.setSelectionMode(QListWidget.NoSelection)
        self.pred_param_list.setStyleSheet(f"""
            QListWidget {{
                background: {COLORS['bg_input']};
                border: 1px solid {COLORS['border']};
                border-radius: 4px;
            }}
            QListWidget::item {{ color: {COLORS['text_primary']}; spacing: 4px; }}
        """)
        self.pred_param_list.itemChanged.connect(self._on_compare_params_changed)
        compare_layout.addWidget(self.pred_param_list, 1)

        # 全选 / 反选 / 清空
        self.pred_select_all_btn = QPushButton("全选")
        self.pred_invert_btn = QPushButton("反选")
        self.pred_clear_btn = QPushButton("清空")
        for _b in (self.pred_select_all_btn, self.pred_invert_btn, self.pred_clear_btn):
            _b.setFixedHeight(26)
            _b.setStyleSheet(f"""
                QPushButton {{
                    background: {COLORS['bg_input']};
                    color: {COLORS['text_primary']};
                    border: 1px solid {COLORS['border']};
                    border-radius: 4px;
                    font-size: 12px;
                    padding: 0 8px;
                }}
                QPushButton:hover {{ background: {COLORS['secondary']}; }}
            """)
        self.pred_select_all_btn.clicked.connect(self._select_all_params)
        self.pred_invert_btn.clicked.connect(self._invert_params)
        self.pred_clear_btn.clicked.connect(self._clear_params)
        compare_layout.addWidget(self.pred_select_all_btn)
        compare_layout.addWidget(self.pred_invert_btn)
        compare_layout.addWidget(self.pred_clear_btn)

        # 模式开关：同排口多参数 / 同参数多排口
        mode_lbl = QLabel("模式:")
        mode_lbl.setStyleSheet(f"color: {COLORS['text_primary']}; font-size: 12px;")
        compare_layout.addWidget(mode_lbl)
        self.pred_mode_combo = QComboBox()
        self.pred_mode_combo.addItems(["同排口多参数", "同参数多排口"])
        self.pred_mode_combo.setCurrentText(self._pred_chart_mode)
        self.pred_mode_combo.setFixedHeight(26)
        self.pred_mode_combo.setStyleSheet(f"""
            QComboBox {{
                background: {COLORS['bg_input']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 4px;
                padding: 0 6px;
                font-size: 12px;
            }}
        """)
        self.pred_mode_combo.currentTextChanged.connect(self._on_pred_mode_changed)
        compare_layout.addWidget(self.pred_mode_combo)

        # 归一化按钮
        self.pred_normalize_btn = QPushButton("归一化: 关")
        self.pred_normalize_btn.setCheckable(True)
        self.pred_normalize_btn.setFixedHeight(26)
        self.pred_normalize_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['bg_input']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 4px;
                font-size: 12px;
                padding: 0 8px;
            }}
            QPushButton:checked {{
                background: {COLORS['primary']};
                color: white;
                font-weight: bold;
            }}
            QPushButton:hover {{ background: {COLORS['secondary']}; }}
        """)
        self.pred_normalize_btn.toggled.connect(self._on_normalize_toggled)
        compare_layout.addWidget(self.pred_normalize_btn)

        pred_chart_content_layout.addWidget(self.pred_compare_panel)

        chart_group = QGroupBox()
        chart_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['secondary']};
                border-radius: 6px;
                margin-top: 0px;
                padding-top: 6px;
            }}
        """)
        chart_inner = QVBoxLayout()
        chart_inner.setContentsMargins(6, 6, 6, 6)
        self.pred_chart = ChartWidget()
        self.pred_chart.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        chart_inner.addWidget(self.pred_chart)
        chart_group.setLayout(chart_inner)
        pred_chart_content_layout.addWidget(chart_group)

        pred_chart_layout.addWidget(pred_chart_header)
        pred_chart_layout.addWidget(self.pred_chart_content, 1)
        layout.addWidget(self.pred_chart_section, 1)

        widget.setLayout(layout)
        return widget

    def _create_settings_tab(self):
        """创建系统设置标签页"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        # 自动刷新设置
        refresh_group = QGroupBox("⚡ 自动刷新设置")
        refresh_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['secondary']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                color: {COLORS['secondary']};
            }}
        """)
        refresh_layout = QVBoxLayout()
        refresh_layout.setSpacing(15)

        # 刷新时间设置
        time_layout = QHBoxLayout()
        time_layout.addWidget(QLabel("刷新间隔:"))

        self.refresh_spinbox = QSpinBox()
        self.refresh_spinbox.setRange(10, 3600)  # 10秒到1小时
        self.refresh_spinbox.setValue(int(self.refresh_interval / 1000))
        self.refresh_spinbox.setSuffix(" 秒")
        self.refresh_spinbox.setStyleSheet(f"""
            QSpinBox {{
                background: {COLORS['bg_input']};
                color: {COLORS['text_primary']};
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
                padding: 8px;
                font-size: 14px;
                font-weight: bold;
            }}
            QSpinBox::up-button, QSpinBox::down-button {{
                background: {COLORS['primary']};
                border: none;
                border-radius: 4px;
                width: 20px;
            }}
            QSpinBox::up-button:hover, QSpinBox::down-button:hover {{
                background: {COLORS['secondary']};
            }}
        """)
        time_layout.addWidget(self.refresh_spinbox)

        time_layout.addStretch()

        apply_refresh_btn = QPushButton("应用设置")
        apply_refresh_btn.clicked.connect(self._apply_refresh_settings)
        apply_refresh_btn.setMinimumWidth(120)
        time_layout.addWidget(apply_refresh_btn)

        refresh_layout.addLayout(time_layout)

        # 说明文本
        info_label = QLabel("💡 提示: 设置数据自动刷新的时间间隔。")
        info_label.setStyleSheet(f"""
            color: {COLORS['text_secondary']};
            padding: 10px;
            background: {COLORS['bg_input']};
            border-radius: 4px;
            border-left: 3px solid {COLORS['accent']};
        """)
        refresh_layout.addWidget(info_label)

        refresh_group.setLayout(refresh_layout)
        layout.addWidget(refresh_group)

        # 预警阈值设置
        threshold_group = QGroupBox("🚨 预警阈值设置")
        threshold_group.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {COLORS['warning']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                color: {COLORS['warning']};
            }}
        """)
        threshold_layout = QVBoxLayout()
        self.threshold_table = QTableWidget()
        self.threshold_table.setColumnCount(4)
        self.threshold_table.setHorizontalHeaderLabels(["参数名称", "下限", "上限", "操作"])
        self.threshold_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.threshold_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.threshold_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.threshold_table.setColumnWidth(3, 80)
        self.threshold_table.setStyleSheet(f"""
            QTableWidget {{
                background: {COLORS['bg_input']};
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
            }}
        """)
        threshold_layout.addWidget(self.threshold_table)
        threshold_group.setLayout(threshold_layout)
        layout.addWidget(threshold_group)

        # 保存按钮和预警历史按钮
        save_layout = QHBoxLayout()
        save_layout.addStretch()

        save_btn = QPushButton("💾 保存所有设置")
        save_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Weight.Bold))
        save_btn.setMinimumWidth(200)
        save_btn.setMinimumHeight(45)
        save_btn.clicked.connect(self._save_thresholds)
        save_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['primary']}, stop:1 {COLORS['secondary']});
                color: white;
                border: 2px solid {COLORS['secondary']};
                border-radius: 10px;
                padding: 12px 30px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['secondary']}, stop:1 {COLORS['primary']});
            }}
        """)
        save_layout.addWidget(save_btn)

        history_btn = QPushButton("📋 查看预警历史")
        history_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Weight.Bold))
        history_btn.setMinimumWidth(200)
        history_btn.setMinimumHeight(45)
        history_btn.clicked.connect(self._show_warning_history)
        history_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['warning']}, stop:1 {COLORS['primary']});
                color: white;
                border: 2px solid {COLORS['primary']};
                border-radius: 10px;
                padding: 12px 30px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['primary']}, stop:1 {COLORS['warning']});
            }}
        """)
        save_layout.addWidget(history_btn)

        layout.addLayout(save_layout)

        # 今日预警统计
        stats_layout = QHBoxLayout()
        stats_layout.addStretch()

        today_count = self.warning_system.get_warning_count_today()
        self.today_warning_label = QLabel(f"📊 今日预警次数: {today_count}")
        self.today_warning_label.setStyleSheet(f"""
            color: {COLORS['text_primary']};
            font-size: 14px;
            padding: 10px 20px;
            background: {COLORS['bg_input']};
            border-radius: 6px;
            border: 1px solid {COLORS['border']};
        """)
        stats_layout.addWidget(self.today_warning_label)

        # 保存标签引用以便后续更新
        self._today_warning_label = self.today_warning_label

        layout.addLayout(stats_layout)

        layout.addStretch()

        self._load_thresholds()
        widget.setLayout(layout)
        return widget

    def _apply_refresh_settings(self):
        """应用刷新设置"""
        interval_seconds = self.refresh_spinbox.value()
        self.refresh_interval = interval_seconds * 1000

        # 重启定时器
        self.refresh_timer.stop()
        self.refresh_timer.start(self.refresh_interval)

        self._update_status_bar()
        QMessageBox.information(self, "设置成功",
                               f"自动刷新间隔已更新为 {interval_seconds} 秒")

    def _check_memory(self):
        """检查并管理内存使用"""
        import psutil
        try:
            process = psutil.Process()
            memory_info = process.memory_info()
            memory_mb = memory_info.rss / 1024 / 1024

            if memory_mb > self._memory_threshold_mb:
                logger.warning(f"[MEMORY] 内存使用较高: {memory_mb:.1f}MB，执行GC")
                gc.collect()

            # 清理过期缓存
            self._cleanup_expired_cache()

            # 记录内存使用（仅在debug模式）
            if memory_mb > 100:
                logger.debug(f"[MEMORY] 当前内存: {memory_mb:.1f}MB")
        except ImportError:
            # psutil 未安装，静默跳过
            pass
        except Exception as e:
            logger.debug(f"[MEMORY] 内存检查失败: {e}")

    def _cleanup_expired_cache(self):
        """清理过期的缓存数据"""
        now = time.time()
        # 清理API缓存
        expired_keys = [
            k for k, v in self._api_data_cache.items()
            if now - v[0] > self._api_cache_ttl
        ]
        for k in expired_keys:
            self._api_data_cache.pop(k, None)

        # 限制API缓存大小
        if len(self._api_data_cache) > 500:
            sorted_keys = sorted(self._api_data_cache, key=lambda k: self._api_data_cache[k][0])
            for k in sorted_keys[:len(sorted_keys) // 3]:
                self._api_data_cache.pop(k, None)

        # 清理预测结果缓存
        if len(self._pred_result_cache) > self._pred_cache_max:
            sorted_keys = sorted(self._pred_result_cache, key=lambda k: self._pred_result_cache[k].get('_time', 0))
            for k in sorted_keys[:len(sorted_keys) // 3]:
                self._pred_result_cache.pop(k, None)

    def _update_status_bar(self):
        """更新状态栏信息"""
        interval_seconds = int(self.refresh_interval / 1000)
        self.statusBar().showMessage(
            f"✓ 已登录: {self.username} | "
            f"📊 版本 {self.version} | "
            f"🔄 自动刷新: {interval_seconds}秒 | "
            f"📡 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )

    def _setup_warning(self):
        """设置预警系统回调"""
        self.warning_system.set_warning_callback(self._on_warning_callback)

    def _on_warning_callback(self, warnings_detail, is_warning, max_level="正常"):
        """预警回调（支持多级预警）"""
        # 更新今日预警次数
        today_count = self.warning_system.get_warning_count_today()
        if hasattr(self, '_today_warning_label'):
            self._today_warning_label.setText(f"📊 今日预警次数: {today_count}")

        if is_warning:
            # 根据最高预警等级设置颜色
            if max_level == "红色预警":
                border_color = "#dc2626"  # 红色
                text_color = "#dc2626"
                icon = "🔴"
            elif max_level == "橙色预警":
                border_color = "#ea580c"  # 橙色
                text_color = "#ea580c"
                icon = "🟠"
            else:  # 黄色预警
                border_color = "#ca8a04"  # 黄色
                text_color = "#ca8a04"
                icon = "🟡"

            text = f"{icon} {max_level}触发:\n\n"
            for w in warnings_detail:
                level_emoji = {
                    "红色预警": "🔴",
                    "橙色预警": "🟠",
                    "黄色预警": "🟡"
                }.get(w.get('level', ''), "")

                wtype = w.get('type', 'realtime')
                if wtype == 'realtime':
                    type_tag = "[实时]"
                elif 'prediction' in wtype:
                    pred_label = wtype.replace('prediction_', '')
                    type_tag = f"[预测·{pred_label}]"
                else:
                    type_tag = ""

                text += f"  {level_emoji} {w.get('ent_name', '')} - {w['subname']}  {type_tag}\n"
                text += f"      {w['param']}: {w['value']} (阈值: {w['threshold']})\n"
                text += f"      等级: {w['level']} | {w.get('datetime', '')}\n\n"

            self.warning_list.setText(text)
            self.warning_list.setStyleSheet(f"""
                color: {text_color};
                padding: 15px;
                background: {COLORS['bg_dark']};
                border-radius: 6px;
                border: 3px solid {border_color};
                font-size: 14px;
                font-weight: bold;
            """)
        else:
            self.warning_list.setText("✅ 暂无预警")
            self.warning_list.setStyleSheet(f"""
                color: {COLORS['success']};
                padding: 15px;
                background: {COLORS['bg_dark']};
                border-radius: 6px;
                border: 1px solid {COLORS['border']};
                font-size: 14px;
            """)

    def _toggle_left_panel(self):
        """切换左侧企业列表的显示/隐藏"""
        if self.left_panel.isVisible():
            self.left_panel.hide()
            self.toggle_btn.setText("▶")
            self.toggle_btn.setToolTip("点击展开企业列表")
        else:
            self.left_panel.show()
            self.toggle_btn.setText("◀")
            self.toggle_btn.setToolTip("点击收起企业列表")

    def _toggle_history_table(self):
        """切换历史数据表格的显示/隐藏"""
        if self.table_content.isVisible():
            # 收起：隐藏内容，将整个 section 压缩到只剩标题栏
            self.table_content.hide()
            self.table_section.setMaximumHeight(40)
            self.table_section.setSizePolicy(
                QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            self.table_toggle_btn.setText("▼ 展开")
        else:
            # 展开：显示内容，解除高度限制，恢复弹性撑满
            self.table_content.show()
            self.table_section.setMaximumHeight(16777215)  # QWIDGETSIZE_MAX
            self.table_section.setSizePolicy(
                QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
            self.table_toggle_btn.setText("▲ 收起")

    def _toggle_history_chart(self):
        """切换历史数据趋势图的显示/隐藏"""
        if self.chart_content.isVisible():
            # 收起：隐藏内容，将整个 section 压缩到只剩标题栏
            self.chart_content.hide()
            self.chart_section.setMaximumHeight(40)
            self.chart_section.setSizePolicy(
                QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            self.chart_toggle_btn.setText("▼ 展开")
        else:
            # 展开：显示内容，解除高度限制，恢复弹性撑满
            self.chart_content.show()
            self.chart_section.setMaximumHeight(16777215)  # QWIDGETSIZE_MAX
            self.chart_section.setSizePolicy(
                QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
            self.chart_toggle_btn.setText("▲ 收起")

    def _toggle_pred_result(self):
        """切换预测结果表格的显示/隐藏"""
        if self.pred_result_content.isVisible():
            # 收起
            self.pred_result_content.hide()
            self.pred_result_section.setMaximumHeight(40)
            self.pred_result_section.setSizePolicy(
                QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            self.pred_result_toggle_btn.setText("▼ 展开")
        else:
            # 展开
            self.pred_result_content.show()
            self.pred_result_section.setMaximumHeight(16777215)
            self.pred_result_section.setSizePolicy(
                QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
            self.pred_result_toggle_btn.setText("▲ 收起")

    def _toggle_pred_chart(self):
        """切换预测趋势图的显示/隐藏"""
        if self.pred_chart_content.isVisible():
            # 收起
            self.pred_chart_content.hide()
            self.pred_chart_section.setMaximumHeight(40)
            self.pred_chart_section.setSizePolicy(
                QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            self.pred_chart_toggle_btn.setText("▼ 展开")
        else:
            # 展开
            self.pred_chart_content.show()
            self.pred_chart_section.setMaximumHeight(16777215)
            self.pred_chart_section.setSizePolicy(
                QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
            self.pred_chart_toggle_btn.setText("▲ 收起")

    def _on_tab_changed(self, index):
        """tab切换事件处理，确保图表正确显示"""
        # 切换到历史数据tab时，重新渲染图表
        if index == 1:  # 历史数据tab
            all_rows = getattr(self, '_history_all_rows', [])
            display_params = getattr(self, '_history_display_params', [])
            codes = getattr(self, '_history_codes', '')
            if all_rows and display_params and codes:
                try:
                    self._display_history_chart(all_rows, display_params, codes)
                except Exception as e:
                    print(f"[ERROR] 切换tab时显示曲线失败: {e}")

        # 切换到预测tab时，如果有预测数据则重新显示
        elif index == 2:  # 预测tab
            if hasattr(self, 'pred_chart') and self.pred_chart:
                # 触发图表重绘
                self.pred_chart.update()

    def _refresh_data(self):
        """刷新实时数据（全局异常保护，防止未捕获异常导致QTimer停止或程序崩溃）"""
        try:
            self._do_refresh_data()
        except Exception as e:
            logger.error(f"刷新数据异常: {e}", exc_info=True)
            # 不弹窗，只记录日志，避免干扰用户；QTimer会继续触发下次刷新
        except BaseException as e:
            # 捕获 MemoryError、SystemError 等非 Exception 子类
            logger.critical(f"刷新数据严重异常(BaseException): {e}", exc_info=True)
            # 尝试GC释放内存
            try:
                gc.collect()
            except Exception:
                pass

    def _schedule_refresh(self):
        """刷新统一入口（三项优化 T01 占位：直接转发到受保护的刷新流程）。

        T04 将在此方法内加入 250ms 节流合并，再调用 ``_flush_refresh()`` →
        ``_do_refresh_data()``。当前实现保持与原有 refresh_timer → _refresh_data 一致的行为，
        是后续所有手动/定时器刷新的唯一入口。
        """
        # 本批直接转发；T04 会改为节流后调用 self._flush_refresh()
        self._refresh_data()

    def _do_refresh_data(self):
        """实际的数据刷新逻辑（从_refresh_data拆出）"""
        self.last_update_label.setText(f"最后更新: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self._update_status_bar()

        # ── 内存管理：定期执行垃圾回收 ───────────────────────────────────────
        self._gc_counter += 1
        if self._gc_counter >= self._gc_interval:
            self._gc_counter = 0
            self._check_memory()

        # 获取所有企业实时数据
        realtime_data = self.multi_client.get_all_realtime_data()
        if not realtime_data.get('rows') or len(realtime_data.get('rows', [])) == 0:
            self.last_update_label.setText(f"最后更新: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} (数据获取失败)")
            logger.warning("获取实时数据失败或无数据")
            # 不弹 QMessageBox，改为状态栏提示，避免阻塞定时器
            return

        # 解析并按排放口分组
        grouped_data = parse_realtime_data(realtime_data)

        # ── 热电厂折算数据注入（预警系统和实时监控都使用） ──────────────────
        # 热电厂（企业名含"热电"）需要使用折算值进行预警判断和实时显示
        # 对 grouped_data 中的热电厂，用折算数据替换实测值
        grouped_data_for_warning = merge_corrected_data_into_grouped(
            grouped_data, self.multi_client, None
        )
        # 实时监控和预警系统都使用包含折算值的数据
        self.current_grouped_data = grouped_data_for_warning

        # 更新排放口列表
        self._update_sub_list(grouped_data)

        # 检查预警（使用包含折算值的数据）
        self.warning_system.check_and_alert(grouped_data_for_warning, self.thresholds)

        # 如果有选中的排放口，显示其数据
        if hasattr(self, 'selected_subid'):
            self._show_sub_realtime(self.selected_subid)

        # 自动刷新预测数据
        self._auto_refresh_prediction()

        logger.info(f"数据刷新完成，共 {len(grouped_data)} 个排放口")

    def _update_sub_list(self, grouped_data):
        """更新排放口列表（支持多企业）"""
        self.sub_list.setRowCount(len(grouped_data))
        for idx, (unique_key, data) in enumerate(grouped_data.items()):
            ent_name = data.get('ent_name', '')
            subname = data.get('subname', '')
            status = data.get('status', '未知')

            # 显示格式：企业名称 - 排口名称
            display_text = f"{ent_name} - {subname}" if ent_name else subname
            self.sub_list.setItem(idx, 0, QTableWidgetItem(display_text))
            status_item = QTableWidgetItem(status)
            status_item.setForeground(QColor(COLORS['success']) if status == '正常' else QColor(COLORS['warning']))
            self.sub_list.setItem(idx, 1, status_item)

            # 同时更新历史查询的下拉框
            exists = self.history_sub_combo.findData(unique_key)
            if exists == -1:
                self.history_sub_combo.addItem(display_text, unique_key)

    def _on_sub_selected(self, item):
        """排放口选择事件"""
        row = item.row()
        text_item = self.sub_list.item(row, 0)
        if text_item:
            display_text = text_item.text()
            # 从grouped_data中找到对应的unique_key
            for unique_key, data in getattr(self, 'current_grouped_data', {}).items():
                ent_name = data.get('ent_name', '')
                subname = data.get('subname', '')
                expected_text = f"{ent_name} - {subname}" if ent_name else subname
                if expected_text == display_text:
                    self.selected_subid = data.get('subid', unique_key)
                    self.selected_unique_key = unique_key
                    self.selected_subtype = data.get('subtype', '')
                    self.selected_ent_name = ent_name
                    self._show_sub_realtime(unique_key)
                    break

    def _show_sub_realtime(self, unique_key):
        """显示指定排放口的实时数据（自动补全实时接口未返回的参数）"""
        if not hasattr(self, 'current_grouped_data'):
            return

        sub_data = self.current_grouped_data.get(unique_key, {})
        params = sub_data.get('params', [])
        sub_type = sub_data.get('subtype', '')
        ent_name = sub_data.get('ent_name', '')
        subname = sub_data.get('subname', '')
        subtype_name = sub_data.get('subtype', '')
        subid = sub_data.get('subid', '')

        # 更新排放口标题
        title = f"📊 {ent_name} - {subname} ({subtype_name})"
        self.sub_info_label.setText(title)

        # ── 第一步：获取该排口真实 itemCode 列表（用于过滤/补全） ────────
        subtype_code = get_subtype_code(sub_type)
        try:
            codes_str = self._get_sub_itemcodes(subid, subname, subtype_code)
        except Exception:
            codes_str = ''
        code_list = [c.strip() for c in codes_str.split(',') if c.strip()] if codes_str else []

        # 构建 code <-> 中文名 映射（只基于该排口的 itemCode）
        code_to_name_local = {code: CODE_TO_NAME.get(code, code) for code in code_list}
        # 中文名 -> code（反向）
        name_to_code_local = {}
        for code, cname in code_to_name_local.items():
            # 优先保留第一个映射（避免多对一覆盖）
            if cname not in name_to_code_local:
                name_to_code_local[cname] = code

        # ── 第二步：确定该排口"应当显示"的参数列表 ──────────────────────
        # 从 GAS_PARAMS_DISPLAY / WATER_PARAMS_DISPLAY 中，只保留该排口 itemCode 里
        # 实际存在的参数（code_to_name_local 的值域），避免显示无关参数
        base_display_params = get_display_params_for_sub(sub_type)
        if code_list:
            # 只保留"在本排口 itemCode 里有对应 code"的参数
            # 同时也保留实时接口已返回的参数（name 包含 base_display_params 中某项）
            itemcode_names = set(code_to_name_local.values())  # 本排口所有参数的中文名
            display_params = [
                d for d in base_display_params
                if d in itemcode_names or any(d in n for n in itemcode_names)
            ]
            if not display_params:
                display_params = base_display_params  # 回退到全量（itemCode获取失败时）
        else:
            display_params = base_display_params

        # ── 第三步：从实时数据筛选已返回的参数（排除"监测中"等无效字符串值） ──
        # 实时值为字符串（如"监测中"）时视为无效，不显示实时值但保留该行占位
        filtered_params = []
        for p in params:
            pname = p.get('name', '')
            if not any(d in pname for d in display_params):
                continue
            val = p.get('value')
            # 若值是字符串且不是数字，视为无效（如"监测中"）
            if isinstance(val, str):
                try:
                    float(val)
                except (ValueError, TypeError):
                    val = None  # 标记为无效，后续用历史补充
            p_copy = dict(p)
            p_copy['value'] = val
            filtered_params.append(p_copy)

        # ── 第四步：找出实时数据中未覆盖的参数，用历史数据补充 ────────────
        covered_display = set()
        for p in filtered_params:
            pname = p.get('name', '')
            if p.get('value') is not None:  # 只有值有效才算已覆盖
                for d in display_params:
                    if d in pname:
                        covered_display.add(d)

        missing_display = [d for d in display_params if d not in covered_display]

        # ── 第三步（补充）：热电厂废气获取折算数据 ──────────────────────
        is_thermal_power = '热电' in ent_name and not is_water_sub(sub_type)
        corrected_values = {}  # code -> 折算值
        corrected_standards = {}  # code -> 折算排放标准（用于预警判断）

        # 打印调试信息
        logger.debug(f"[REALTIME] {ent_name}/{subname}: sub_type='{sub_type}', is_water={is_water_sub(sub_type)}, is_thermal={is_thermal_power}")
        logger.debug(f"[REALTIME] {ent_name}/{subname}: codes_str='{codes_str}', code_list={code_list}")
        logger.debug(f"[REALTIME] {ent_name}/{subname}: filtered_params codes={[(p.get('code'), type(p.get('code')).__name__) for p in filtered_params]}")

        # 如果 codes_str 为空，从实时数据直接提取 code（备用方案）
        if not codes_str and is_thermal_power:
            codes_from_realtime = set()
            for p in params:
                c = p.get('code')
                if c is not None:
                    try:
                        codes_from_realtime.add(str(int(float(c))))
                    except (ValueError, TypeError):
                        codes_from_realtime.add(str(c))
            codes_str = ','.join(sorted(codes_from_realtime, key=lambda x: int(x) if x.isdigit() else 0))
            code_list = list(codes_from_realtime)
            logger.debug(f"[REALTIME] {ent_name}/{subname}: 从实时数据提取codes_str='{codes_str}'")

        if is_thermal_power and codes_str:
            try:
                client_info = self.multi_client.get_client_by_subid(subid)
                if client_info:
                    _client = client_info["client"]
                    # 获取折算数据（showUpload=1）
                    logger.debug(f"[REALTIME] {ent_name}/{subname}: 正在获取折算数据, codes={codes_str}")
                    corrected_result = _client.get_minute_data_current_hour(subid, subtype_code, codes_str, True)
                    logger.debug(f"[REALTIME] {ent_name}/{subname}: API返回={corrected_result.get('rows', [])[:1] if corrected_result.get('rows') else '空'}")
                    corrected_rows = corrected_result.get('rows', [])
                    if corrected_rows:
                        # 取最新一条数据
                        latest = corrected_rows[0]
                        logger.debug(f"[REALTIME] {ent_name}/{subname}: 折算原始数据keys={list(latest.keys())}")
                        # 打印第一条数据的部分值
                        for k, v in list(latest.items())[:10]:
                            logger.debug(f"[REALTIME] {ent_name}/{subname}:   {k} = {v}")
                        for code in code_list:
                            # 尝试多种可能的折算字段名
                            found = False
                            for key in [f'cvt_{code}', f'Corrected_{code}', f'val_{code}', 'Cvt', 'Corrected', 'cvt']:
                                if key in latest:
                                    cvt_val = latest.get(key)
                                    if cvt_val is not None and cvt_val != '':
                                        try:
                                            corrected_values[code] = float(cvt_val)
                                            logger.debug(f"[REALTIME] {ent_name}/{subname}: code={code}, key={key}, val={cvt_val}")
                                            found = True
                                        except (ValueError, TypeError):
                                            pass
                                        break
                            if not found:
                                logger.debug(f"[REALTIME] {ent_name}/{subname}: code={code} 未找到折算字段")
                            # 同时获取折算排放标准
                            std_key = f'stand_{code}'
                            if std_key in latest:
                                std_val = latest.get(std_key)
                                if std_val is not None and std_val != '' and std_val != '-9999':
                                    try:
                                        corrected_standards[code] = float(std_val)
                                        logger.debug(f"[REALTIME] {ent_name}/{subname}: code={code}, 折算标准={std_val}")
                                    except (ValueError, TypeError):
                                        pass
                        logger.debug(f"[REALTIME] {ent_name}/{subname}: 获取到 {len(corrected_values)} 个折算值, {len(corrected_standards)} 个折算标准")
                        logger.debug(f"[REALTIME] {ent_name}/{subname}: corrected_values keys={list(corrected_values.keys())}")
                    else:
                        logger.debug(f"[REALTIME] {ent_name}/{subname}: 折算数据为空，error={corrected_result.get('error', '无')}")
            except Exception as e:
                logger.warning(f"[WARN] 获取热电厂折算数据失败: {e}")

        if missing_display and codes_str:
            # 查最近2小时历史（分钟数据 index=-1），取最新有效分钟值
            hist_latest = {}
            hist_latest_time = {}
            try:
                from datetime import datetime as _dt, timedelta as _td
                client_info = self.multi_client.get_client_by_subid(subid)
                if client_info:
                    _client = client_info["client"]
                    _end   = _dt.now()
                    _start = _end - _td(hours=2)
                    _result = _client.query_history(
                        subid, subtype_code, codes_str,
                        _start.strftime('%Y-%m-%d %H:%M'),
                        _end.strftime('%Y-%m-%d %H:%M'),
                        index=-1, page=1, rows=120
                    )
                    hist_rows = _result.get('rows', [])
                    if hist_rows:
                        # 从最新到最旧逐行取值，取第一个有值的
                        _hist_time_raw = ''
                        for h_row in reversed(hist_rows):
                            if not _hist_time_raw:
                                _tr = h_row.get('DateTime', '')
                                if _tr:
                                    # 格式化时间：取 HH:MM（去掉秒和日期）
                                    try:
                                        if ' ' in _tr:
                                            _tpart = _tr.split(' ')[-1]
                                        else:
                                            _tpart = _tr
                                        _parts = _tpart.split(':')
                                        _hist_time_raw = f"{_parts[0]}:{_parts[1]}" if len(_parts) >= 2 else _tr
                                    except Exception:
                                        _hist_time_raw = _tr
                            for code in code_list:
                                vk = f"val_{code}"
                                hv = h_row.get(vk)
                                if hv is not None and code not in hist_latest:
                                    hist_latest[code] = hv
                                    hist_latest_time[code] = _hist_time_raw
            except Exception as e:
                print(f"[WARN] 补充历史数据失败: {e}")

            # 将缺失参数补充进 filtered_params（注意：code不在itemCode里的跳过，不显示）
            for missing_name in missing_display:
                code = name_to_code_local.get(missing_name)
                if not code:
                    # 该参数不在本排口的 itemCode 里 → 跳过，不显示
                    continue

                hist_val = hist_latest.get(code)
                # 从实时数据的 params 里找标准值
                std_val = ''
                for p in params:
                    if p.get('code') == code:
                        std_val = p.get('std', '')
                        break
                    # 也尝试通过 name 匹配
                    if missing_name in p.get('name', ''):
                        std_val = p.get('std', '')
                        break

                filtered_params.append({
                    'name': missing_name,
                    'value': hist_val,
                    'std': std_val,
                    'is_exceed': False,
                    '_is_hist': hist_val is not None,
                    '_is_补充': True,
                    '_hist_time': hist_latest_time.get(code, ''),
                })

        # ── 渲染表格 ────────────────────────────────────────────────────
        # 列顺序: 0=监测参数, 1=实测浓度, 2=折算浓度, 3=单位, 4=排放标准, 5=折算状态, 6=监测状态
        self.realtime_table.setRowCount(len(filtered_params))
        for row, param in enumerate(filtered_params):
            name       = param.get('name', '')
            code_raw   = param.get('code', '')
            # 统一转换为纯数字字符串格式（去掉 .0 后缀）
            try:
                code = str(int(float(code_raw)))
            except (ValueError, TypeError):
                code = str(code_raw).strip()
            value      = param.get('value')       # 实测值
            corrected  = corrected_values.get(code)  # 折算值
            std        = param.get('std', '')
            is_exceed  = param.get('is_exceed', False)
            is_hist    = param.get('_is_hist', False)    # 来自历史数据
            is_补充    = param.get('_is_补充', False)    # 实时接口未返回

            self.realtime_table.setItem(row, 0, QTableWidgetItem(name))

            # ── 实测浓度列 ──────────────────────────────────────────────
            if value is not None:
                val_text = str(value) + ("  (历史)" if is_hist else "")
            else:
                val_text = "--"
            value_item = QTableWidgetItem(val_text)
            if is_exceed:
                value_item.setForeground(QColor(COLORS['warning']))
            elif is_hist:
                value_item.setForeground(QColor("#aaaaaa"))
            self.realtime_table.setItem(row, 1, value_item)

            # ── 折算浓度列（热电厂废气显示，其他留空）──────────────────
            if is_thermal_power:
                if corrected is not None:
                    cvt_text = f"{corrected:.2f}"
                else:
                    cvt_text = "--"
                cvt_item = QTableWidgetItem(cvt_text)
                # 折算值超标时标红（优先使用折算标准判断）
                try:
                    # 优先使用服务端返回的折算标准
                    cvt_std = corrected_standards.get(code, std)
                    if cvt_std:
                        cvt_std_val = float(cvt_std) if isinstance(cvt_std, str) else cvt_std
                        if cvt_std_val > 0 and corrected is not None and corrected > cvt_std_val:
                            cvt_item.setForeground(QColor(COLORS['warning']))
                    elif std:
                        std_val = float(std) if std else 0
                        if std_val > 0 and corrected is not None and corrected > std_val:
                            cvt_item.setForeground(QColor(COLORS['warning']))
                except (ValueError, TypeError):
                    pass
                self.realtime_table.setItem(row, 2, cvt_item)
            else:
                self.realtime_table.setItem(row, 2, QTableWidgetItem("--"))

            # ── 单位列 ─────────────────────────────────────────────────
            self.realtime_table.setItem(row, 3, QTableWidgetItem(""))

            # ── 排放标准列 ─────────────────────────────────────────────
            # 热电厂废气显示折算标准，其他显示实测标准
            if is_thermal_power and corrected_standards.get(code):
                display_std = corrected_standards.get(code, std)
            else:
                display_std = std
            self.realtime_table.setItem(row, 4, QTableWidgetItem(str(display_std) if display_std else str(std)))

            # ── 折算状态列（热电厂废气显示折算是否超标）────────────────
            if is_thermal_power:
                if corrected is not None:
                    try:
                        # 优先使用服务端返回的折算标准
                        cvt_std = corrected_standards.get(code, std)
                        if cvt_std:
                            cvt_std_val = float(cvt_std) if isinstance(cvt_std, str) else cvt_std
                            if cvt_std_val > 0:
                                cvt_exceed = corrected > cvt_std_val
                                cvt_status_text = "超标" if cvt_exceed else "正常"
                                cvt_status_color = QColor(COLORS['warning']) if cvt_exceed else QColor(COLORS['success'])
                            else:
                                cvt_status_text = "正常"
                                cvt_status_color = QColor(COLORS['success'])
                        elif std:
                            std_val = float(std)
                            if std_val > 0:
                                cvt_exceed = corrected > std_val
                                cvt_status_text = "超标" if cvt_exceed else "正常"
                                cvt_status_color = QColor(COLORS['warning']) if cvt_exceed else QColor(COLORS['success'])
                            else:
                                cvt_status_text = "正常"
                                cvt_status_color = QColor(COLORS['success'])
                        else:
                            cvt_status_text = "正常"
                            cvt_status_color = QColor(COLORS['success'])
                    except (ValueError, TypeError):
                        cvt_status_text = "--"
                        cvt_status_color = QColor("#888888")
                else:
                    cvt_status_text = "--"
                    cvt_status_color = QColor("#888888")
                cvt_status_item = QTableWidgetItem(cvt_status_text)
                cvt_status_item.setForeground(cvt_status_color)
                self.realtime_table.setItem(row, 5, cvt_status_item)
            else:
                self.realtime_table.setItem(row, 5, QTableWidgetItem("--"))

            # ── 监测状态列 ─────────────────────────────────────────────
            # 热电厂废气：使用折算值判断是否超标；其他：使用实测值判断
            if is_补充 and value is None:
                status_text = "暂无数据"
                status_color = QColor("#888888")
            elif is_hist:
                _ht = param.get('_hist_time', '')
                status_text = f"{_ht}历史" if _ht else "历史参考"
                status_color = QColor("#aaaaaa")
            else:
                # 热电厂废气：优先用折算值判断，其次用实测值
                if is_thermal_power:
                    thermal_exceed = False
                    if corrected is not None:
                        try:
                            cvt_std = corrected_standards.get(code, std)
                            std_val = float(cvt_std) if cvt_std else (float(std) if std else 0)
                            thermal_exceed = (std_val > 0 and corrected > std_val)
                        except (ValueError, TypeError):
                            thermal_exceed = is_exceed
                    else:
                        thermal_exceed = is_exceed
                    status_text = "超标" if thermal_exceed else "正常"
                    status_color = QColor(COLORS['warning']) if thermal_exceed else QColor(COLORS['success'])
                else:
                    status_text = "超标" if is_exceed else "正常"
                    status_color = QColor(COLORS['warning']) if is_exceed else QColor(COLORS['success'])
            status_item = QTableWidgetItem(status_text)
            status_item.setForeground(status_color)
            self.realtime_table.setItem(row, 6, status_item)

    def _on_time_type_changed(self, index):
        """数据类型切换时，智能显示/隐藏时分控件，并调整默认时间范围"""
        # 0=分钟, 1=小时, 2=日, 3=月, 4=季度, 5=年
        need_time = index in (0, 1)   # 分钟和小时需要时分控件
        self.start_time.setVisible(need_time)
        self.end_time.setVisible(need_time)

        today = QDate.currentDate()
        if index == 0:    # 分钟 → 默认今天
            self.start_date.setDate(today)
            self.end_date.setDate(today)
        elif index == 1:  # 小时 → 默认近7天
            self.start_date.setDate(today.addDays(-7))
            self.end_date.setDate(today)
        elif index == 2:  # 日 → 默认近30天
            self.start_date.setDate(today.addDays(-30))
            self.end_date.setDate(today)
        elif index == 3:  # 月 → 默认近6个月
            self.start_date.setDate(today.addMonths(-6))
            self.end_date.setDate(today)
        elif index == 4:  # 季度 → 默认近1年
            self.start_date.setDate(today.addMonths(-12))
            self.end_date.setDate(today)
        elif index == 5:  # 年 → 默认近3年
            self.start_date.setDate(today.addYears(-3))
            self.end_date.setDate(today)

    def _quick_set_time_range(self, key):
        """快捷设置时间范围"""
        today = QDate.currentDate()
        tt_idx = self.time_type_combo.currentIndex()  # 当前数据类型

        if key == "today":
            self.start_date.setDate(today)
            self.end_date.setDate(today)
            self.start_time.setTime(QTime(0, 0))
            self.end_time.setTime(QTime(23, 59))
        elif key == "yesterday":
            yesterday = today.addDays(-1)
            self.start_date.setDate(yesterday)
            self.end_date.setDate(yesterday)
            self.start_time.setTime(QTime(0, 0))
            self.end_time.setTime(QTime(23, 59))
        elif key == "7d":
            self.start_date.setDate(today.addDays(-6))
            self.end_date.setDate(today)
        elif key == "30d":
            self.start_date.setDate(today.addDays(-29))
            self.end_date.setDate(today)
        elif key == "this_month":
            self.start_date.setDate(QDate(today.year(), today.month(), 1))
            self.end_date.setDate(today)
            # 若是月/季度/年数据，自动切换类型
            if tt_idx <= 1:  # 分钟/小时 → 切日数据
                self.time_type_combo.setCurrentIndex(2)
        elif key == "last_month":
            first_this = QDate(today.year(), today.month(), 1)
            last_month_end = first_this.addDays(-1)
            last_month_start = QDate(last_month_end.year(), last_month_end.month(), 1)
            self.start_date.setDate(last_month_start)
            self.end_date.setDate(last_month_end)
            if tt_idx <= 1:
                self.time_type_combo.setCurrentIndex(2)
        elif key == "this_quarter":
            month = today.month()
            quarter_start_month = ((month - 1) // 3) * 3 + 1
            self.start_date.setDate(QDate(today.year(), quarter_start_month, 1))
            self.end_date.setDate(today)
            self.time_type_combo.setCurrentIndex(4)   # 切到季度数据
        elif key == "this_year":
            self.start_date.setDate(QDate(today.year(), 1, 1))
            self.end_date.setDate(today)
            self.time_type_combo.setCurrentIndex(5)   # 切到年数据
        elif key == "last_year":
            last_year = today.year() - 1
            self.start_date.setDate(QDate(last_year, 1, 1))
            self.end_date.setDate(QDate(last_year, 12, 31))
            self.time_type_combo.setCurrentIndex(5)

    def _query_history(self):
        """查询历史数据（支持分钟/小时/日/月/季度/年）"""
        unique_key = self.history_sub_combo.currentData()
        if not unique_key:
            QMessageBox.warning(self, "提示", "请选择排放口")
            return

        tt_idx = self.time_type_combo.currentIndex()
        # index: -1=分钟, 1=小时, 2=日, 3=月  （API 不直接支持季度/年，用月数据后端聚合）
        index_map = {0: -1, 1: 1, 2: 2, 3: 3, 4: 3, 5: 3}
        index = index_map.get(tt_idx, 1)
        time_type_label = ["分钟数据", "小时数据", "日数据", "月数据", "季度数据", "年数据"][tt_idx]

        # v5.18: 日/月数据服务端不支持(index=2/3返回空)，改为查小时数据后客户端聚合
        _original_index = index
        if index >= 2:
            index = 1  # 用小时数据
            print(f"[DEBUG] 日/月查询改为小时数据聚合: original_index={_original_index}")

        # 构造时间字符串
        start_date = self.start_date.date().toString("yyyy-MM-dd")
        end_date   = self.end_date.date().toString("yyyy-MM-dd")

        if tt_idx in (0, 1):  # 分钟/小时：带时分
            start_time = self.start_time.time().toString("HH:mm")
            end_time   = self.end_time.time().toString("HH:mm")
            start = f"{start_date} {start_time}"
            end   = f"{end_date} {end_time}"
        else:  # 日/月/季度/年：只用日期
            start = f"{start_date} 00:00"
            end   = f"{end_date} 23:59"

        # 获取排放口信息
        sub_data = self.current_grouped_data.get(unique_key, {})
        sub_type = sub_data.get('subtype', '')
        subid = sub_data.get('subid', unique_key)
        subtype_code = get_subtype_code(sub_type)
        subname = sub_data.get('subname', '')
        ent_name = sub_data.get('ent_name', '')

        # 判断是否为热电厂废气
        is_thermal_power = '热电' in ent_name and not is_water_sub(sub_type)

        # 获取监测项目代码 - 使用GetSubs接口获取itemCode
        try:
            codes = self._get_sub_itemcodes(subid, subname, subtype_code)
        except Exception as e:
            QMessageBox.warning(self, "查询失败", f"获取监测项目代码失败:\n{str(e)}")
            return

        if not codes:
            QMessageBox.warning(self, "提示", "无法获取监测项目代码")
            return

        # 获取显示参数列表（用于表格展示）
        display_params = get_display_params_for_sub(sub_type)

        # 查询（自动分页获取全量数据）
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            all_rows = []
            all_corrected_rows = []  # 折算数据
            current_page = 1
            page_size = 1000  # 每页请求1000条
            total = 0

            while True:
                # 请求当前页
                try:
                    result = self.multi_client.query_history(
                        subid, subtype_code, codes, start, end, index,
                        page=current_page, rows=page_size
                    )
                    # 如果是热电厂废气，同时获取折算数据
                    if is_thermal_power:
                        result_corrected = self.multi_client.query_history(
                            subid, subtype_code, codes, start, end, index,
                            page=current_page, rows=page_size, use_corrected=True
                        )
                except Exception as e:
                    QMessageBox.warning(self, "查询失败", f"第 {current_page} 页请求失败:\n{str(e)}")
                    return

                # 检查返回结果格式
                if not isinstance(result, dict):
                    QMessageBox.warning(self, "查询失败", f"第 {current_page} 页返回格式错误:\n{str(result)[:200]}")
                    return

                page_rows = result.get('rows', [])
                page_total = result.get('total', 0)

                # 收集折算数据
                if is_thermal_power and 'result_corrected' in dir():
                    page_corrected = result_corrected.get('rows', [])
                    all_corrected_rows.extend(page_corrected)

                # 第一页时记录 total
                if current_page == 1:
                    total = page_total

                # 打印调试信息
                print(f"[DEBUG] 第 {current_page} 页: 返回 {len(page_rows)} 条, total={page_total}")

                # 累加数据
                all_rows.extend(page_rows)

                # 判断是否还有下一页
                # 条件1: 本次返回的条数少于请求的条数（说明是最后一页）
                # 条件2: 已获取的条数 >= total（说明已经全部获取完毕）
                if len(page_rows) < page_size or len(all_rows) >= total:
                    break

                current_page += 1
                # 最多请求10页，防止无限循环
                if current_page > 10:
                    print(f"[DEBUG] 已请求10页，停止分页")
                    break

            # 打印最终统计
            print(f"[DEBUG] 历史数据查询完成: start={start}, end={end}, index={index}")
            print(f"[DEBUG] 服务端 total={total}, 实际获取={len(all_rows)} 条")

            # v5.18: 日/月数据客户端聚合（从小时数据聚合）
            if _original_index >= 2 and all_rows:
                all_rows = self._aggregate_hour_to_day_month(all_rows, codes, _original_index)
                total = len(all_rows)
                print(f"[DEBUG] 聚合后 index={_original_index} 共 {len(all_rows)} 条")
            # 日/月/年查询调试：打印第一行数据的DateTime和前3个字段
            if all_rows:
                _sample = all_rows[0]
                _dt = _sample.get('DateTime', 'N/A')
                _keys = [k for k in _sample.keys() if not k.startswith('stand_') and not k.startswith('state_')][:5]
                print(f"[DEBUG] 首行时间={_dt}, 字段={_keys}")
            else:
                print(f"[DEBUG] 未返回任何数据行，请检查时间范围或index参数")

            # ── 季度/年数据：将月数据在客户端聚合 ────────────────────────
            if tt_idx in (4, 5) and all_rows:
                all_rows = self._aggregate_to_quarter_or_year(all_rows, codes, tt_idx)
                print(f"[DEBUG] 聚合后 {time_type_label} 共 {len(all_rows)} 条")

            # ── 热电厂废气：将折算数据合并到主数据 ───────────────────────
            if is_thermal_power and all_corrected_rows:
                print(f"[DEBUG] 合并折算数据: {len(all_corrected_rows)} 条")
                # 按时间戳建立折算数据索引
                corrected_index = {}
                for cr in all_corrected_rows:
                    dt = cr.get('DateTime', '')
                    if dt:
                        corrected_index[dt] = cr
                # 将折算值添加到主数据行
                for row in all_rows:
                    dt = row.get('DateTime', '')
                    if dt and dt in corrected_index:
                        cr = corrected_index[dt]
                        # 为每个code添加折算字段
                        for code in codes.split(','):
                            code = code.strip()
                            if code:
                                cvt_key = f'cvt_{code}'
                                if cvt_key in cr:
                                    row[cvt_key] = cr[cvt_key]

            # 保存全量数据，供分页和导出使用
            self._history_all_rows = all_rows
            self._history_display_params = display_params
            self._history_codes = codes
            self._history_time_type_label = time_type_label   # 供导出使用
            self._history_is_thermal_power = is_thermal_power   # 标记是否热电厂废气

            # 分页参数
            self._history_page_size = 50
            self._history_current_page = 1

            # 显示第一页
            self._render_history_page()

            # 显示曲线（全量数据）
            if all_rows:
                try:
                    self._display_history_chart(all_rows, display_params, codes)
                except Exception as e:
                    print(f"[ERROR] 显示曲线失败: {e}")
                    # 曲线显示失败不影响表格，只打印日志

            # 提示用户查询结果
            msg = f"查询完成！\n\n数据类型: {time_type_label}\n时间范围: {start} ~ {end}\n获取数据: {len(all_rows)} 条"
            if total > 0 and total != len(all_rows) and tt_idx not in (4, 5):
                msg += f"\n服务端总数: {total} 条"
            if len(all_rows) == 0:
                QMessageBox.information(self, "查询结果", msg)

        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"[ERROR] 历史数据查询异常:\n{error_detail}")
            QMessageBox.critical(self, "查询失败", f"查询历史数据时发生错误:\n{str(e)}")
        finally:
            QApplication.restoreOverrideCursor()

    def _aggregate_to_quarter_or_year(self, rows, codes_str, tt_idx):
        """将月粒度数据聚合为季度/年粒度（取均值），返回聚合后的行列表"""
        from collections import defaultdict
        import re

        code_list = [c.strip() for c in codes_str.split(',') if c.strip()]

        def get_period_key(dt_str):
            """根据 DateTime 字段提取分组 key"""
            # DateTime 格式通常是 "2025-03" 或 "2025-03-01 00:00"
            m = re.match(r'(\d{4})-(\d{2})', dt_str or '')
            if not m:
                return None
            year, month = int(m.group(1)), int(m.group(2))
            if tt_idx == 4:  # 季度
                q = (month - 1) // 3 + 1
                return f"{year}年 Q{q}"
            else:  # 年
                return f"{year}年"

        # 按 period_key 分组
        groups = defaultdict(list)
        for row in rows:
            key = get_period_key(row.get('DateTime', ''))
            if key:
                groups[key].append(row)

        # 对每组计算均值
        aggregated = []
        for period_key in sorted(groups.keys()):
            group_rows = groups[period_key]
            agg_row = {'DateTime': period_key}
            for code in code_list:
                vk = f"val_{code}"
                vals = [float(r[vk]) for r in group_rows if r.get(vk) not in (None, '', '--')]
                agg_row[vk] = round(sum(vals) / len(vals), 4) if vals else None
            aggregated.append(agg_row)

        return aggregated

    def _aggregate_hour_to_day_month(self, rows, codes_str, target_index):
        """将小时数据聚合为日(index=2)/月(index=3)数据，取均值"""
        from collections import defaultdict
        import re

        code_list = [c.strip() for c in codes_str.split(',') if c.strip()]

        def get_group_key(dt_str):
            """从 DateTime 提取分组 key"""
            dt_str = dt_str or ''
            if target_index == 2:  # 日数据：按 yyyy-MM-dd 分组
                m = re.match(r'(\d{4}-\d{2}-\d{2})', dt_str)
                return m.group(1) if m else None
            else:  # 月数据：按 yyyy-MM 分组
                m = re.match(r'(\d{4}-\d{2})', dt_str)
                return m.group(1) if m else None

        # 按日期/月份分组
        groups = defaultdict(list)
        for row in rows:
            key = get_group_key(row.get('DateTime', ''))
            if key:
                groups[key].append(row)

        # 对每组计算均值
        aggregated = []
        for group_key in sorted(groups.keys()):
            group_rows = groups[group_key]
            agg_row = {'DateTime': group_key}
            for code in code_list:
                vk = f"val_{code}"
                vals = []
                for r in group_rows:
                    v = r.get(vk)
                    if v is not None and v != '' and v != '--':
                        try:
                            vals.append(float(v))
                        except (ValueError, TypeError):
                            pass
                agg_row[vk] = round(sum(vals) / len(vals), 4) if vals else None
                # 保留标准值和状态（取第一行的）
                sk = f"stand_{code}"
                if sk in group_rows[0]:
                    agg_row[sk] = group_rows[0].get(sk)
                stk = f"state_{code}"
                if stk in group_rows[0]:
                    agg_row[stk] = group_rows[0].get(stk)
            aggregated.append(agg_row)

        return aggregated

    def _render_history_page(self):
        """按当前页码渲染历史数据表格，并更新分页控件状态"""
        rows = getattr(self, '_history_all_rows', [])
        display_params = getattr(self, '_history_display_params', [])
        codes = getattr(self, '_history_codes', '')
        page_size = getattr(self, '_history_page_size', 50)
        current_page = getattr(self, '_history_current_page', 1)

        total = len(rows)
        total_pages = max(1, (total + page_size - 1) // page_size)
        current_page = max(1, min(current_page, total_pages))
        self._history_current_page = current_page

        start_idx = (current_page - 1) * page_size
        end_idx = min(start_idx + page_size, total)
        page_rows = rows[start_idx:end_idx]

        self._display_history_table(page_rows, display_params, codes)

        # 更新分页控件
        self.page_info_label.setText(
            f"第 {current_page} 页 / 共 {total_pages} 页  (共 {total} 条)"
        )
        self.page_prev_btn.setEnabled(current_page > 1)
        self.page_next_btn.setEnabled(current_page < total_pages)

    def _history_prev_page(self):
        """上一页"""
        self._history_current_page = getattr(self, '_history_current_page', 1) - 1
        self._render_history_page()

    def _history_next_page(self):
        """下一页"""
        self._history_current_page = getattr(self, '_history_current_page', 1) + 1
        self._render_history_page()

    def _get_sub_itemcodes(self, subid: str, subname: str, subtype_code: str) -> str:
        """
        获取排口的监测项目代码列表
        首先尝试从缓存获取，如果没有则调用API获取
        如果API返回的itemCode为空，则从实时数据中提取C0001_ITEM_CODE
        """
        # 检查缓存
        cache_key = f"{subid}_{subtype_code}"
        if hasattr(self, '_sub_itemcodes_cache') and cache_key in self._sub_itemcodes_cache:
            return self._sub_itemcodes_cache[cache_key]

        # 根据subid找到对应的client和username
        client_info = self.multi_client.get_client_by_subid(subid)
        if not client_info:
            print(f"[WARNING] 无法根据subid找到对应的client: subid={subid}")
            return ""
        client = client_info["client"]
        username = client_info["username"]

        # 调用API获取企业列表
        ent_list = client.get_enterprise_list(subtype_code)
        if not ent_list:
            print(f"[WARNING] 无法获取企业列表, 尝试从实时数据提取itemCode: subtype_code={subtype_code}")
            # 备用方案: 从实时数据中提取itemCode
            return self._extract_itemcodes_from_realtime(client, subid, cache_key)

        # 找到当前登录的企业ID
        ent_id = None
        for ent in ent_list:
            ent_name = ent.get('name', '')
            # 尝试多种匹配方式:
            # 1. 完全匹配
            if ent_name == username or ent_name == client.enterprise_name:
                ent_id = ent.get('id')
                print(f"[DEBUG] 企业完全匹配: username={username}, ent_name={ent_name}")
                break
            # 2. 包含匹配（处理API返回的企业名称与配置不完全一致的情况）
            if username in ent_name or ent_name in username:
                ent_id = ent.get('id')
                print(f"[DEBUG] 企业包含匹配: username={username}, ent_name={ent_name}")
                break
            # 3. 去空格后匹配
            if ent_name.replace(' ', '') == username.replace(' ', '') or \
               ent_name.replace(' ', '') == client.enterprise_name.replace(' ', ''):
                ent_id = ent.get('id')
                print(f"[DEBUG] 企业去空格匹配: username={username}, ent_name={ent_name}")
                break

        if not ent_id:
            # 如果没找到，打印调试信息
            print(f"[WARNING] 未找到匹配的企业, username={username}, enterprise_name={client.enterprise_name}")
            print(f"[DEBUG] 企业列表: {[(e.get('name'), e.get('id')) for e in ent_list]}")
            # 尝试使用第一个企业
            if ent_list:
                ent_id = ent_list[0].get('id')
                print(f"[DEBUG] 使用第一个企业: {ent_list[0].get('name')}, id={ent_id}")

        if not ent_id:
            print(f"[WARNING] 无法获取企业ID, 尝试从实时数据提取itemCode")
            return self._extract_itemcodes_from_realtime(client, subid, cache_key)

        # 获取排口列表
        subs = client.get_sub_list(ent_id, subtype_code)
        if not subs:
            print(f"[WARNING] 无法获取排口列表, 尝试从实时数据提取itemCode")
            return self._extract_itemcodes_from_realtime(client, subid, cache_key)

        # 找到对应的排口并获取itemCode
        # 注意：subid可能是浮点数（如1716.0），需要转换为整数比较
        try:
            subid_int = int(float(subid))
        except (ValueError, TypeError):
            subid_int = None

        for sub in subs:
            sub_info_id = sub.get('id', '')
            sub_info_name = sub.get('name', '')

            # 尝试通过ID匹配
            match_by_id = False
            if subid_int is not None:
                try:
                    sub_info_id_int = int(float(sub_info_id))
                    if sub_info_id_int == subid_int:
                        match_by_id = True
                except (ValueError, TypeError):
                    pass

            # 尝试通过名称匹配
            match_by_name = sub_info_name == subname

            if match_by_id or match_by_name:
                itemcode = sub.get('itemCode', '')
                if itemcode:
                    # 缓存结果
                    if not hasattr(self, '_sub_itemcodes_cache'):
                        self._sub_itemcodes_cache = {}
                    self._sub_itemcodes_cache[cache_key] = itemcode
                    print(f"[DEBUG] 从排口列表获取itemCode: {itemcode}")
                    return itemcode
                else:
                    print(f"[WARNING] 排口itemCode为空, 尝试从实时数据提取: {subname}")
                    # 备用方案: 从实时数据中提取itemCode
                    return self._extract_itemcodes_from_realtime(client, subid, cache_key)

        print(f"[WARNING] 未找到匹配的排口, 尝试从实时数据提取itemCode")
        # 备用方案: 从实时数据中提取itemCode
        return self._extract_itemcodes_from_realtime(client, subid, cache_key)

    def _extract_itemcodes_from_realtime(self, client, subid: str, cache_key: str) -> str:
        """
        从实时数据中提取监测项目代码（备用方案）
        当API返回的itemCode为空时使用
        """
        try:
            # 获取实时数据
            realtime_resp = client.get_realtime_data()
            realtime_data = realtime_resp.get('rows', []) if isinstance(realtime_resp, dict) else []

            if not realtime_data:
                print(f"[ERROR] 实时数据为空, 无法提取itemCode")
                return ""

            # 转换subid为整数比较
            try:
                subid_int = int(float(subid))
            except (ValueError, TypeError):
                print(f"[ERROR] 无法转换subid为整数: {subid}")
                return ""

            # 筛选当前排口的数据
            sub_data = []
            for row in realtime_data:
                row_subid = row.get('C0007_SUBSTATION_ID', '')
                try:
                    row_subid_int = int(float(row_subid))
                    if row_subid_int == subid_int:
                        sub_data.append(row)
                except (ValueError, TypeError):
                    pass

            if not sub_data:
                print(f"[ERROR] 未找到排口 {subid} 的实时数据")
                return ""

            # 提取所有不同的C0001_ITEM_CODE
            itemcodes = set()
            for row in sub_data:
                item_code = row.get('C0001_ITEM_CODE')
                if item_code is not None:
                    # 转换为整数去重
                    try:
                        itemcode_int = int(float(item_code))
                        itemcodes.add(str(itemcode_int))
                    except (ValueError, TypeError):
                        pass

            if itemcodes:
                # 排序并转换为逗号分隔的字符串
                itemcode_str = ','.join(sorted(itemcodes, key=lambda x: int(x) if x.isdigit() else 0))
                # 缓存结果
                if not hasattr(self, '_sub_itemcodes_cache'):
                    self._sub_itemcodes_cache = {}
                self._sub_itemcodes_cache[cache_key] = itemcode_str
                print(f"[DEBUG] 从实时数据提取itemCode: {itemcode_str} (共{len(itemcodes)}个监测项目)")
                return itemcode_str
            else:
                print(f"[ERROR] 实时数据中没有找到C0001_ITEM_CODE字段")
                return ""

        except Exception as e:
            print(f"[ERROR] 从实时数据提取itemCode时发生异常: {str(e)}")
            return ""

    def _display_history_table(self, rows, display_params, codes_str):
        """显示历史数据表格"""
        # 解析codes字符串，提取代码列表
        code_list = [c.strip() for c in codes_str.split(',') if c.strip()]

        # 构建代码到名称的映射
        from data_processor import CODE_TO_NAME
        code_to_display = {code: CODE_TO_NAME.get(code, code) for code in code_list}

        # 检查是否为热电厂废气
        is_thermal_power = getattr(self, '_history_is_thermal_power', False)

        # 只显示有代码的参数
        display_params_filtered = [code_to_display.get(code, code) for code in code_list if code]

        if is_thermal_power:
            # 热电厂废气：时间 + 实测1 + 折算1 + 实测2 + 折算2 + ...
            # 每个参数占2列
            total_cols = 1 + len(display_params_filtered) * 2
            headers = ["时间"]
            for param in display_params_filtered:
                headers.append(f"{param}(实测)")
                headers.append(f"{param}(折算)")
        else:
            total_cols = len(display_params_filtered) + 1
            headers = ["时间"] + display_params_filtered

        self.history_table.setColumnCount(total_cols)
        self.history_table.setHorizontalHeaderLabels(headers)

        # 设置列宽：时间列固定宽度，其他列自动拉伸
        self.history_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.history_table.setColumnWidth(0, 150)  # 时间列固定150像素
        for i in range(1, total_cols):
            self.history_table.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)

        self.history_table.setRowCount(len(rows))
        for row_idx, row in enumerate(rows):
            self.history_table.setItem(row_idx, 0, QTableWidgetItem(row.get('DateTime', '')))
            if is_thermal_power:
                # 热电厂废气：每个参数2列（实测 + 折算）
                for col_idx, code in enumerate(code_list):
                    base_col = 1 + col_idx * 2
                    val_key = f"val_{code}"
                    cvt_key = f"cvt_{code}"
                    val = row.get(val_key)
                    cvt = row.get(cvt_key)
                    self.history_table.setItem(row_idx, base_col, QTableWidgetItem(str(val) if val else "--"))
                    self.history_table.setItem(row_idx, base_col + 1, QTableWidgetItem(str(cvt) if cvt else "--"))
            else:
                for col_idx, code in enumerate(code_list, 1):
                    val_key = f"val_{code}"
                    val = row.get(val_key)
                    self.history_table.setItem(row_idx, col_idx, QTableWidgetItem(str(val) if val else "--"))


    def _display_history_chart(self, rows, display_params, codes_str):
        """显示历史数据曲线图"""
        # 解析codes字符串
        code_list = [c.strip() for c in codes_str.split(',') if c.strip()]

        # 构建代码到名称的映射
        from data_processor import CODE_TO_NAME
        code_to_display = {code: CODE_TO_NAME.get(code, code) for code in code_list}

        # 检查是否为热电厂废气
        is_thermal_power = getattr(self, '_history_is_thermal_power', False)

        times = [r.get('DateTime', '') for r in rows]

        series_list = []
        for code in code_list:
            # 实测曲线
            values = []
            for row in rows:
                val_key = f"val_{code}"
                val = row.get(val_key)
                values.append(float(val) if val else None)
            series_list.append({"name": code_to_display.get(code, code) + "(实测)", "data": values})

            # 折算曲线（仅热电厂废气）
            if is_thermal_power:
                cvt_values = []
                for row in rows:
                    cvt_key = f"cvt_{code}"
                    cvt = row.get(cvt_key)
                    cvt_values.append(float(cvt) if cvt else None)
                series_list.append({"name": code_to_display.get(code, code) + "(折算)", "data": cvt_values})

        self.chart_widget.plot_series(times, series_list, title="历史数据曲线")

    def _export_history(self):
        """导出历史数据到 Excel（带数据类型标识，完整格式）"""
        all_rows = getattr(self, '_history_all_rows', None)
        if not all_rows:
            QMessageBox.warning(self, "提示", "暂无数据，请先查询历史数据")
            return

        # 构建默认文件名
        time_type_label = getattr(self, '_history_time_type_label', '历史数据')
        subname = ""
        unique_key = self.history_sub_combo.currentData()
        if unique_key and hasattr(self, 'current_grouped_data'):
            sd = self.current_grouped_data.get(unique_key, {})
            ent  = sd.get('ent_name', '')
            name = sd.get('subname', '')
            subname = f"{ent}-{name}" if ent else name

        start_str = self.start_date.date().toString("yyyyMMdd")
        end_str   = self.end_date.date().toString("yyyyMMdd")
        default_name = f"{subname}_{time_type_label}_{start_str}~{end_str}.xlsx" if subname \
                       else f"{time_type_label}_{start_str}~{end_str}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", default_name, "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            import pandas as pd
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            codes_str = getattr(self, '_history_codes', '')
            code_list = [c.strip() for c in codes_str.split(',') if c.strip()]
            from data_processor import CODE_TO_NAME
            code_to_display = {code: CODE_TO_NAME.get(code, code) for code in code_list}

            # 构建数据记录
            records = []
            for row in all_rows:
                record = {"时间": row.get('DateTime', '')}
                for code in code_list:
                    val = row.get(f"val_{code}")
                    record[code_to_display.get(code, code)] = val if val is not None else ""
                records.append(record)

            df = pd.DataFrame(records)

            # 写入 Excel 并美化格式
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=time_type_label[:31], index=False, startrow=2)
                ws = writer.sheets[time_type_label[:31]]

                # ── 第1行：大标题 ────────────────────────────────────────
                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=max(1, len(df.columns)))
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = (
                    f"{subname}  {time_type_label}  "
                    f"{self.start_date.date().toString('yyyy-MM-dd')} ~ "
                    f"{self.end_date.date().toString('yyyy-MM-dd')}"
                )
                title_cell.font      = Font(bold=True, size=13, color="FFFFFF")
                title_cell.fill      = PatternFill("solid", fgColor="2C3E50")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 28

                # ── 第3行：表头美化 ─────────────────────────────────────
                header_fill   = PatternFill("solid", fgColor="3498DB")
                header_font   = Font(bold=True, color="FFFFFF", size=11)
                header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thin_border   = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin')
                )
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=3, column=col_idx)
                    cell.font      = header_font
                    cell.fill      = header_fill
                    cell.alignment = header_align
                    cell.border    = thin_border

                # ── 数据行：交替背景色 + 边框 ─────────────────────────
                fill_even = PatternFill("solid", fgColor="EAF4FB")
                fill_odd  = PatternFill("solid", fgColor="FFFFFF")
                data_align = Alignment(horizontal="center", vertical="center")
                for row_idx in range(4, 4 + len(records)):
                    fill = fill_even if (row_idx % 2 == 0) else fill_odd
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill      = fill
                        cell.alignment = data_align
                        cell.border    = thin_border

                # ── 自动列宽 ────────────────────────────────────────────
                for col_idx, col_name in enumerate(df.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = max(
                        len(str(col_name)),
                        *[len(str(r.get(col_name, ''))) for r in records[:50]]
                    ) if records else len(str(col_name))
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

            QMessageBox.information(
                self, "导出成功",
                f"已导出 {len(records)} 条 {time_type_label} 到:\n{file_path}"
            )
        except ImportError as e:
            QMessageBox.critical(self, "导出失败", f"缺少依赖模块: {str(e)}\n请联系管理员")
        except PermissionError:
            QMessageBox.warning(self, "导出失败", f"文件被占用，请关闭后重试:\n{file_path}")
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.warning(self, "导出失败", f"导出时发生错误:\n{str(e)}")

    def _on_pred_type_changed(self, index):
        """预测类型改变 —— 清空旧结果"""
        self.pred_table.setRowCount(0)
        self.pred_result_label.setText('点击"开始预测"查看预测结果（将对所有排放口的所有参数进行预测）')
        self.pred_chart.clear()

    def _on_sound_alarm_changed(self, state):
        """声音报警开关状态改变"""
        self.sound_alarm_enabled = (state == 2)
        self.warning_system.set_sound_enabled(self.sound_alarm_enabled)
        status = "已开启 🔔" if self.sound_alarm_enabled else "已关闭 🔕"
        print(f"[INFO] 声音报警 {status}")
        # 若当前已有预测结果，更新底部提示文字
        if self.pred_table.rowCount() > 0:
            old_text = self.pred_result_label.text()
            # 替换声音状态片段
            import re as _re
            old_text = _re.sub(r'声音报警：.*', f'声音报警：{status}', old_text)
            self.pred_result_label.setText(old_text)

    def _on_sound_type_changed(self, index):
        """声音类型下拉框改变"""
        type_map = {0: 'beep1', 1: 'beep2', 2: 'beep3'}
        self.sound_alarm_type = type_map.get(index, 'beep1')
        self.warning_system.set_sound_type(self.sound_alarm_type)
        print(f"[INFO] 声音类型切换为: {self.sound_alarm_type}")

    def _on_pred_horizon_changed(self, index):
        """预测未来小时数改变（1/2/3小时）"""
        self.prediction_horizon = index + 1   # index 0→1h, 1→2h, 2→3h
        # 清空表格，提示用户重新预测
        self.pred_table.setRowCount(0)
        self.pred_result_label.setText(
            f'已切换为预测未来 {self.prediction_horizon} 小时，请点击"开始预测"更新结果')
        print(f"[INFO] 预测未来小时数: {self.prediction_horizon}h")

    def _show_prediction_params_dialog(self):
        """显示预测指标选择对话框"""
        import traceback
        
        try:
            print("[DEBUG] 开始显示预测指标对话框")
            # CODE_TO_NAME 已在文件顶部导入，直接使用
            if 'CODE_TO_NAME' not in globals():
                QMessageBox.warning(self, "错误", "CODE_TO_NAME 未定义")
                return
            
            # 获取所有可用参数
            all_params = list(CODE_TO_NAME.values())
            all_params.sort()  # 按字母排序
            print(f"[DEBUG] 加载了 {len(all_params)} 个参数")
        except Exception as e:
            error_msg = f"无法加载参数列表: {str(e)}\n{traceback.format_exc()}"
            print(f"[ERROR] {error_msg}")
            QMessageBox.warning(self, "错误", f"无法加载参数列表: {str(e)}")
            return
        
        try:
            # 创建对话框
            dialog = QDialog(self)
            dialog.setWindowTitle("选择预测指标")
            dialog.setMinimumWidth(400)
            dialog.setMinimumHeight(500)
            
            layout = QVBoxLayout()
            layout.setSpacing(15)
            layout.setContentsMargins(20, 20, 20, 20)
            
            # 标题
            title = QLabel("请选择需要预测的指标：")
            title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50;")
            layout.addWidget(title)
            
            # 提示
            hint = QLabel("不勾选任何指标 = 预测所有指标")
            hint.setStyleSheet("color: #7f8c8d; font-size: 12px;")
            layout.addWidget(hint)
            
            # 复选框列表（使用滚动区域）
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setMinimumHeight(350)
            
            checkbox_widget = QWidget()
            checkbox_layout = QVBoxLayout(checkbox_widget)
            checkbox_layout.setSpacing(8)
            
            checkboxes = {}
            for param in all_params:
                try:
                    cb = QCheckBox(param)
                    cb.setStyleSheet("""
                        QCheckBox {
                            font-size: 14px;
                            padding: 5px;
                        }
                        QCheckBox::indicator {
                            width: 18px;
                            height: 18px;
                        }
                    """)
                    # 如果该参数已在预测列表中，默认选中
                    if param in self.prediction_params:
                        cb.setChecked(True)
                    checkbox_layout.addWidget(cb)
                    checkboxes[param] = cb
                except Exception as e:
                    print(f"[ERROR] 创建复选框失败 {param}: {e}")
                    continue
            
            checkbox_layout.addStretch()
            scroll.setWidget(checkbox_widget)
            layout.addWidget(scroll)
            
            # 按钮区域
            btn_layout = QHBoxLayout()
            btn_layout.addStretch()
            
            select_all_btn = QPushButton("全选")
            select_all_btn.setMinimumWidth(80)
            select_all_btn.setStyleSheet("""
                QPushButton {
                    background: #3498db;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 16px;
                    font-size: 13px;
                }
                QPushButton:hover {
                    background: #2980b9;
                }
            """)
            select_all_btn.clicked.connect(lambda: [cb.setChecked(True) for cb in checkboxes.values()])
            btn_layout.addWidget(select_all_btn)
            
            clear_btn = QPushButton("清空")
            clear_btn.setMinimumWidth(80)
            clear_btn.setStyleSheet("""
                QPushButton {
                    background: #95a5a6;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 16px;
                    font-size: 13px;
                }
                QPushButton:hover {
                    background: #7f8c8d;
                }
            """)
            clear_btn.clicked.connect(lambda: [cb.setChecked(False) for cb in checkboxes.values()])
            btn_layout.addWidget(clear_btn)
            
            btn_layout.addSpacing(10)
            
            ok_btn = QPushButton("确定")
            ok_btn.setMinimumWidth(80)
            ok_btn.setStyleSheet("""
                QPushButton {
                    background: #27ae60;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 16px;
                    font-size: 13px;
                }
                QPushButton:hover {
                    background: #229954;
                }
            """)
            btn_layout.addWidget(ok_btn)
            
            cancel_btn = QPushButton("取消")
            cancel_btn.setMinimumWidth(80)
            cancel_btn.setStyleSheet("""
                QPushButton {
                    background: #e74c3c;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 16px;
                    font-size: 13px;
                }
                QPushButton:hover {
                    background: #c0392b;
                }
            """)
            btn_layout.addWidget(cancel_btn)
            
            layout.addLayout(btn_layout)
            
            dialog.setLayout(layout)
            
            # 按钮事件
            ok_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)
            
            print("[DEBUG] 显示对话框")
            # 显示对话框
            result = dialog.exec_()
            print(f"[DEBUG] 对话框结果: {result}")
            
            if result == QDialog.Accepted:
                print("[DEBUG] 用户点击确定")
                # 获取选中的参数
                selected = [param for param, cb in checkboxes.items() if cb.isChecked()]
                
                # 更新预测参数列表
                self.prediction_params = selected
                
                # 显示结果
                if not selected:
                    msg = "已设置为预测所有指标"
                    self.pred_result_label.setText(
                        f'点击"开始预测"查看预测结果（将对所有排放口的所有参数进行预测）')
                else:
                    msg = f"已选择 {len(selected)} 个预测指标: {', '.join(selected[:5])}"
                    if len(selected) > 5:
                        msg += f" 等..."
                    self.pred_result_label.setText(
                        f'已选择 {len(selected)} 个预测指标，点击"开始预测"查看结果')
                
                print(f"[INFO] {msg}")
                
                # 清空旧结果，让用户重新预测
                self.pred_table.setRowCount(0)
                self.pred_chart.clear()
                
                # 自动重新预测（使用用户选择的参数）
                if hasattr(self, 'current_grouped_data') and self.current_grouped_data:
                    self._run_prediction(silent=True, predict_all_params=False)
                    
        except Exception as e:
            error_msg = f"预测指标对话框出错: {str(e)}\n{traceback.format_exc()}"
            print(f"[ERROR] {error_msg}")
            QMessageBox.critical(self, "错误", f"预测指标对话框出错:\n{str(e)}\n\n请查看控制台获取详细信息。")

    def _on_auto_update_changed(self, state):
        """自动更新状态改变"""
        self.auto_update_prediction = (state == 2)  # 2表示选中
        if self.auto_update_prediction:
            self.prediction_timer.start(self.prediction_update_interval)
            print("预测自动更新已开启")
        else:
            self.prediction_timer.stop()
            print("预测自动更新已暂停")

    def _auto_update_prediction(self):
        """自动更新预测数据"""
        if not self.auto_update_prediction:
            return
        try:
            self._run_prediction(silent=True)
        except Exception as e:
            print(f"自动更新预测失败: {e}")

    def _run_prediction(self, silent=False, predict_all_params=True):
        """运行预测 —— 遍历所有排放口，预测全量参数（异步模式，API调用在后台线程）
        
        Args:
            silent: 是否静默模式（不显示等待光标）
            predict_all_params: 是否预测所有参数（True=预测所有，False=使用用户选择的参数）
        """
        if not hasattr(self, 'current_grouped_data') or not self.current_grouped_data:
            if not silent:
                QMessageBox.warning(self, "提示", "暂无排放口数据，请先刷新数据")
            return

        # 防止重复启动：上一个 worker 尚未结束时跳过
        if hasattr(self, '_pred_worker') and self._pred_worker is not None:
            try:
                if self._pred_worker.isRunning():
                    print("[PRED] 上一次预测尚未完成，跳过本次请求")
                    return
            except RuntimeError:
                # worker 已被删除，重置引用
                self._pred_worker = None

        pred_type = self.pred_type_combo.currentText()
        # 显示等待光标（主线程立即返回，不阻塞）
        if not silent:
            QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)

        # 默认预测所有参数（不传prediction_params或传空列表）
        # 只有从预测指标对话框点击确定时才使用用户选择的参数
        params_to_predict = [] if predict_all_params else self.prediction_params
        print(f"[PRED] 启动预测，类型: {pred_type}, 预测参数: {'全部' if predict_all_params else params_to_predict}")

        # 启动后台线程执行预测
        print(f"[PRED] 启动预测，缓存状态: {hasattr(self, '_sub_itemcodes_cache')} keys={list(self._sub_itemcodes_cache.keys()) if hasattr(self, '_sub_itemcodes_cache') else 'N/A'}")
        self._pred_worker = PredictionWorker(
            grouped_data=self.current_grouped_data,
            multi_client=self.multi_client,
            prediction_horizon=self.prediction_horizon,
            prediction_params=params_to_predict,
            pred_type=pred_type,
            thresholds=self.thresholds,
            sub_itemcodes_cache=self._sub_itemcodes_cache,
            api_data_cache=self._api_data_cache,
            thresholds_cache=getattr(self, '_thresholds_cache', {}),
            get_subtype_code_fn=get_subtype_code,
            is_water_sub_fn=is_water_sub,
            code_to_name_map=CODE_TO_NAME,
            intervention_sm=getattr(self, '_intervention_sm', None),
            get_sub_itemcodes_fn=self._get_sub_itemcodes,
        )
        self._pred_worker.prediction_done.connect(
            lambda preds: self._on_prediction_worker_done(preds, pred_type, silent)
        )
        self._pred_worker.finished.connect(self._pred_worker.deleteLater)
        self._pred_worker.start()

    def _on_prediction_worker_done(self, all_predictions, pred_type, silent):
        """后台线程完成后在主线程执行UI更新"""
        QApplication.restoreOverrideCursor()

        # 调试日志
        print(f"[PRED] 预测完成，类型: {pred_type}, 结果数量: {len(all_predictions)}")
        if all_predictions:
            print(f"[PRED] 第一条预测: {all_predictions[0]}")
        else:
            print("[PRED] 警告: 预测结果为空！")

        # ── 预测增量检查：对比指纹，有变化才重建UI ─────────────────────────
        def _pred_fingerprint(preds):
            fp = {}
            for p in preds:
                key = (p.get('ent_name', ''), p.get('subname', ''),
                       p.get('param', ''), p.get('pred_type', ''))
                cur_val = p.get('cur_pred') or p.get('predicted')
                fp[key] = round(cur_val, 4) if cur_val is not None else None
            return fp

        cur_fp = _pred_fingerprint(all_predictions)
        prev_fp = getattr(self, '_last_pred_fingerprint', {})
        ui_changed = (cur_fp != prev_fp)
        print(f"[PRED] UI需要更新: {ui_changed}")

        # ── 显示结果 ────────────────────────────────────────────────────────
        if ui_changed:
            if "小时数据预测" in pred_type:
                self._show_hour_prediction_table(all_predictions)
            else:
                self._show_day_prediction_table(all_predictions)
        else:
            print("[PRED] 预测数据无变化，跳过UI重建")
        self._last_pred_fingerprint = cur_fp

        # ── 对预测超标触发声音报警 ──────────────────────────────────────────
        from data_processor import get_warning_level as _gwl
        pred_warn_items = []
        for p in all_predictions:
            future_seq = p.get('future_preds', [])
            param_name = p.get('param', '')
            if future_seq:
                for fh in future_seq:
                    fval = fh.get('predicted')
                    if fval is None:
                        continue
                    flevel = _gwl(fval, param_name, self.thresholds)
                    cap = fh.get('warning_level_cap')
                    if cap is not None:
                        level_order2 = {"正常": 0, "黄色预警": 1, "橙色预警": 2, "红色预警": 3}
                        flevel = cap if level_order2.get(cap, 0) >= level_order2.get(flevel, 0) else flevel
                    if flevel != '正常':
                        pred_warn_items.append({
                            'ent_name':     p.get('ent_name', ''),
                            'subname':      p.get('subname', ''),
                            'param':        param_name,
                            'predicted':    fval,
                            'warning_level': flevel,
                            'pred_type':    fh.get('label', '预测'),
                            'confidence':   fh.get('confidence', 1.0),
                        })
            elif p.get('is_warning'):
                pred_warn_items.append({
                    'ent_name':     p.get('ent_name', ''),
                    'subname':      p.get('subname', ''),
                    'param':        param_name,
                    'predicted':    p.get('predicted'),
                    'warning_level': p.get('warning_level', '正常'),
                    'pred_type':    '日均预测',
                    'confidence':   p.get('confidence', 1.0),
                })

        if pred_warn_items:
            self.warning_system.check_and_alert_predictions(pred_warn_items, self.thresholds)
        else:
            self.warning_system.check_and_alert_predictions([], self.thresholds)

        # 如果开启了自动更新，启动定时器
        if self.auto_update_prediction and not self.prediction_timer.isActive():
            self.prediction_timer.start(self.prediction_update_interval)

    def _get_cached_api_result(self, cache_key: str, ttl: int, fetch_fn, *args, **kwargs):
        """
        API结果缓存：同一 cache_key 在 ttl 秒内返回缓存结果，避免重复请求。
        返回 (rows, from_cache)
        """
        now = time.time()
        entry = self._api_data_cache.get(cache_key)
        if entry is not None:
            ts, data = entry
            if now - ts < ttl:
                return data, True
        # 缓存过期或不存在，重新请求
        data = fetch_fn(*args, **kwargs)
        self._api_data_cache[cache_key] = (now, data)
        # 防止缓存无限膨胀
        if len(self._api_data_cache) > 500:
            # 删除最老的1/3条目
            sorted_keys = sorted(self._api_data_cache, key=lambda k: self._api_data_cache[k][0])
            for k in sorted_keys[:166]:
                self._api_data_cache.pop(k, None)
        return data, False

    def _do_prediction_for_sub(self, client, subid, subtype_code, codes,
                                code_list, code_to_name, pred_type,
                                subname='', ent_name=''):
        """对单个排放口执行预测，返回预测列表"""
        predictions = []
        from data_processor import check_threshold, get_warning_level

        if "小时数据预测" in pred_type:
            # 废水排口（氨氮、COD等）上报频率为小时级，分钟数据无意义；
            # 改用今日小时数据作为预测基础，避免读到流量等无关字段产生虚假趋势。
            is_water = is_water_sub(sub_type)
            if is_water:
                # 废水小时数据：TTL=55秒（更新间隔比刷新周期稍短即可）
                cache_key = f"{subid}_today_hour_{codes}"
                result, _ = self._get_cached_api_result(
                    cache_key, 55,
                    client.get_today_hour_data, subid, subtype_code, codes
                )
            else:
                # 废气分钟数据：TTL=40秒
                # 判断是否为热电厂废气（企业名含"热电"且非废水）
                is_thermal_gas = '热电' in ent_name and not is_water_sub(sub_type)
                cache_key = f"{subid}_minute_{codes}_cvt{int(is_thermal_gas)}"
                result, _ = self._get_cached_api_result(
                    cache_key, 40,
                    client.get_minute_data_current_hour, subid, subtype_code, codes,
                    use_corrected=is_thermal_gas  # 热电厂使用折算数据
                )
            rows = result.get('rows', [])

            horizon = getattr(self, 'prediction_horizon', 3)

            for code in code_list:
                param_name = code_to_name.get(code, code)
                # prediction_params 为空则不过滤，预测全部参数
                if self.prediction_params and param_name not in self.prediction_params:
                    continue

                # 获取值：折算数据优先使用 cvt_{code} 字段
                val_key = f"val_{code}"
                cvt_key = f"cvt_{code}"
                
                # 对于热电厂废气，优先使用折算值
                if is_thermal_gas:
                    raw_values = []
                    for r in rows:
                        v = r.get(cvt_key)
                        if v is None or v == '' or v == 0:
                            v = r.get(val_key)  # 备用使用实测值
                        if v not in (None, '', 0):
                            try:
                                raw_values.append(float(v))
                            except (ValueError, TypeError):
                                pass
                else:
                    raw_values = [float(r.get(val_key)) for r in rows if r.get(val_key) not in (None, '', 0)]
                
                if not raw_values:
                    continue

                # 异常值过滤：用 IQR 方法剔除离群点，防止单个异常值拉偏趋势
                if len(raw_values) >= 4:
                    sorted_v = sorted(raw_values)
                    q1 = sorted_v[len(sorted_v) // 4]
                    q3 = sorted_v[3 * len(sorted_v) // 4]
                    iqr = q3 - q1
                    lo = q1 - 3.0 * iqr
                    hi = q3 + 3.0 * iqr
                    filtered = [v for v in raw_values if lo <= v <= hi]
                    input_values = filtered if filtered else raw_values
                else:
                    input_values = raw_values

                # 废水参数（小时数据）用 predict_day_average 更合适，
                # 废气分钟数据仍用 predict_future_hours
                if is_water:
                    from data_processor import predict_day_average, get_warning_level as _gwl2
                    day_res   = predict_day_average(input_values, with_trend=True)
                    day_pred  = day_res.get('predicted')
                    # 构造与 future_seq 格式相同的结构，方便后续统一处理
                    future_seq = []
                    for offset, label in [(0, '当前小时'), (1, '+1小时'), (2, '+2小时'), (3, '+3小时')][:horizon + 1]:
                        future_seq.append({
                            'hour_offset': offset,
                            'label': label,
                            'predicted': day_pred,
                            'confidence': round(day_res.get('confidence', 0) * (0.85 ** offset), 2),
                            'trend': day_res.get('trend', 'stable'),
                            'trend_rate': 0,
                            'data_points': day_res.get('data_points', 0),
                            'data_completeness': day_res.get('data_completeness', 0),
                        })
                else:
                    # ── 干预状态机判断（替代简单的 already_warned）─────────────────
                    # 根据最近数据趋势自动识别干预状态，而非仅看当前是否超标
                    cur_mean = sum(input_values) / len(input_values) if input_values else 0
                    threshold = self.thresholds.get(param_name, 0) if param_name in self.thresholds else 0

                    # 更新干预状态机，获取干预参数
                    if cur_mean > 0 and threshold > 0:
                        import time as _time
                        intervention_params = self._intervention_sm.update(
                            subname=subname,
                            param=param_name,
                            cur_value=cur_mean,
                            threshold=threshold,
                            current_time=_time.time()
                        )
                    else:
                        intervention_params = None

                    # 预测当前小时 + 未来 horizon 小时（共 horizon+1 条）
                    # 传入阈值参数，使预测上限更合理（超标后不会无限增长）
                    future_seq = predict_future_hours(input_values, horizon=horizon,
                                                      intervention_params=intervention_params,
                                                      threshold=threshold)

                # 取每个时间点的预警等级，求最高
                level_order = {"正常": 0, "黄色预警": 1, "橙色预警": 2, "红色预警": 3}
                max_level   = "正常"
                for fh in future_seq:
                    fval = fh.get('predicted')
                    if fval is None:
                        continue
                    lv = get_warning_level(fval, param_name, self.thresholds)
                    if level_order.get(lv, 0) > level_order.get(max_level, 0):
                        max_level = lv

                # ── 实时锚定约束 ────────────────────────────────────────────────
                # 当前实际数据均值若尚未超标（正常），则预测预警等级上限为橙色。
                # 纯粹靠趋势外推不允许直接触发红色预警，避免从正常跳红造成误报。
                if input_values and param_name in self.thresholds:
                    cur_avg = sum(input_values) / len(input_values)
                    cur_actual_level = get_warning_level(cur_avg, param_name, self.thresholds)
                    if cur_actual_level == "正常" and max_level == "红色预警":
                        max_level = "橙色预警"
                        # 同步压制各时间点的预测等级
                        for fh in future_seq:
                            if fh.get('predicted') is not None:
                                lv = get_warning_level(fh['predicted'], param_name, self.thresholds)
                                if lv == "红色预警":
                                    fh['warning_level_cap'] = "橙色预警"

                is_warning = (max_level != "正常")
                cur_item   = future_seq[0]   # 当前小时
                cur_pred   = cur_item.get('predicted')

                predictions.append({
                    'ent_name':    ent_name,
                    'subname':     subname,
                    'param':       param_name,
                    # 当前小时（兼容旧字段）
                    'cur_pred':    cur_pred,
                    # 完整的未来多小时序列，供表格填充
                    'future_preds': future_seq,        # [{hour_offset, label, predicted, confidence, ...}]
                    'trend':       cur_item.get('trend', 'stable'),
                    'trend_rate':  cur_item.get('trend_rate', 0),
                    'confidence':  cur_item.get('confidence', 0),
                    'data_points': cur_item.get('data_points', 0),
                    'warning_level': max_level,
                    'is_warning':  is_warning,
                    'pred_type':   '小时预测',
                    # 用于预警系统（取当前小时预测值）
                    'predicted':   cur_pred,
                })

        else:  # 当日均值预测
            cache_key = f"{subid}_today_hour_{codes}"
            result, _ = self._get_cached_api_result(
                cache_key, 55,
                client.get_today_hour_data, subid, subtype_code, codes
            )
            rows = result.get('rows', [])

            for code in code_list:
                param_name = code_to_name.get(code, code)
                if self.prediction_params and param_name not in self.prediction_params:
                    continue

                val_key = f"val_{code}"
                hour_values = [float(r.get(val_key)) for r in rows if r.get(val_key)]
                if not hour_values:
                    continue

                pred_result   = predict_day_average(hour_values, with_trend=True)
                predicted     = pred_result.get('predicted')
                warning_level = get_warning_level(predicted, param_name, self.thresholds) if predicted is not None else "正常"
                is_warning    = (warning_level != "正常")

                predictions.append({
                    'ent_name':    ent_name,
                    'subname':     subname,
                    'param':       param_name,
                    'cur_pred':    predicted,
                    'next_pred':   None,   # 日均预测无"下一小时"
                    'trend':       pred_result.get('trend', 'stable'),
                    'trend_rate':  pred_result.get('trend_rate', 0),
                    'confidence':  pred_result.get('confidence', 0),
                    'next_confidence': 0,
                    'data_points': pred_result.get('data_points', 0),
                    'warning_level': warning_level,
                    'is_warning':  is_warning,
                    'pred_type':   '日均预测',
                    'predicted':   predicted,
                    'method':      pred_result.get('method', 'unknown'),
                })

        return predictions

    def _auto_refresh_prediction(self):
        """自动刷新预测数据（切换排放口时触发，静默模式）"""
        try:
            self._run_prediction(silent=True)
        except Exception as e:
            print(f"[DEBUG] 自动刷新预测失败: {e}")

    # ── 预测结果表格显示 ────────────────────────────────────────────────────

    @staticmethod
    def _fmt_pred_val(val):
        """格式化预测数值"""
        if val is None:
            return "—"
        try:
            return f"{float(val):.2f}"
        except (TypeError, ValueError):
            return "—"

    @staticmethod
    def _trend_text(trend, trend_rate):
        """趋势文字"""
        if trend == 'up':
            return f"↑ 上升 {trend_rate:+.1f}%"
        elif trend == 'down':
            return f"↓ 下降 {trend_rate:+.1f}%"
        return "→ 平稳"

    @staticmethod
    def _warning_level_color(level):
        """预警等级 → 背景色"""
        return {
            '黄色预警': '#7a6a00',
            '橙色预警': '#7a3d00',
            '红色预警': '#7a0000',
        }.get(level, None)

    def _rebuild_pred_table_columns(self, horizon: int):
        """
        根据 horizon 动态重建 pred_table 的列结构。
        固定列：企业/排放口 | 参数 | 趋势 | 当前置信度 | 预警状态
        动态列（插在参数之后）：当前小时 | +1小时 | +2小时 | …
        """
        # 固定列：企业/排放口(0)、参数(1)、[动态预测列]、趋势、置信度、预警状态
        pred_labels = ["当前小时"] + [f"+{i}小时" for i in range(1, horizon + 1)]
        headers = ["企业/排放口", "参数"] + pred_labels + ["趋势", "置信度", "预警状态"]
        ncols = len(headers)

        self.pred_table.setColumnCount(ncols)
        self.pred_table.setHorizontalHeaderLabels(headers)

        hh = self.pred_table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for c in range(1, ncols):
            hh.setSectionResizeMode(c, QHeaderView.ResizeMode.ResizeToContents)

        # 应用样式
        from PyQt5.QtGui import QColor
        self.pred_table.setStyleSheet(f"""
            QTableWidget {{
                background: {COLORS['bg_input']};
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
                gridline-color: {COLORS['border']};
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {COLORS['bg_card']};
                color: {COLORS['text_primary']};
                padding: 6px;
                border: 1px solid {COLORS['border']};
                font-weight: bold;
            }}
            QTableWidget::item {{
                padding: 4px 8px;
                color: {COLORS['text_primary']};
            }}
        """)
        return pred_labels   # 返回动态列标签，便于填值时索引

    def _fill_pred_table(self, predictions, horizon: int = 0):
        """
        将预测列表填充到 self.pred_table 表格中。

        horizon > 0 时（小时预测）：动态列 = 当前小时 + 未来 horizon 小时
        horizon == 0 时（日均预测）：仅1列「当日均值预测」

        每行预测数据结构:
          - 小时预测：含 'future_preds' 字段（predict_future_hours 返回）
          - 日均预测：含 'cur_pred' / 'predicted' 字段
        """
        from PyQt5.QtGui import QColor
        from PyQt5.QtWidgets import QTableWidgetItem
        from PyQt5.QtCore import Qt

        self.pred_table.setRowCount(0)

        if not predictions:
            self.pred_result_label.setText("暂无预测结果")
            return

        # ── 1. 构建列头 ──────────────────────────────────────────────────────
        if horizon > 0:
            pred_labels = self._rebuild_pred_table_columns(horizon)
        else:
            # 日均预测：只有1个预测列
            pred_labels = self._rebuild_pred_table_columns(0)
            # 将"当前小时"列头改为"当日均值预测"
            self.pred_table.setHorizontalHeaderItem(
                2, QTableWidgetItem("当日均值预测"))

        total_pred_cols = len(pred_labels)     # 预测值列数
        # 列索引：企业/排放口=0, 参数=1, 预测值[2…2+total_pred_cols-1], 趋势, 置信度, 预警状态
        col_trend  = 2 + total_pred_cols
        col_conf   = col_trend + 1
        col_warn   = col_conf  + 1

        # 预警状态文字映射
        warn_text_map = {
            '正常':     '✅ 正常',
            '黄色预警': '⚠ 黄色预警',
            '橙色预警': '🔶 橙色预警',
            '红色预警': '🚨 红色预警',
        }

        self.pred_table.setRowCount(len(predictions))

        for row_idx, p in enumerate(predictions):
            # ── 列0：企业/排放口 ─────────────────────────────────────────────
            ent = p.get('ent_name', '')
            sub = p.get('subname', '')
            col0_text = f"{ent}\n{sub}" if ent else sub
            item0 = QTableWidgetItem(col0_text)
            item0.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            self.pred_table.setItem(row_idx, 0, item0)

            # ── 列1：参数 ────────────────────────────────────────────────────
            item1 = QTableWidgetItem(p.get('param', ''))
            item1.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.pred_table.setItem(row_idx, 1, item1)

            # ── 动态预测值列（列2 … 2+total_pred_cols-1）────────────────────
            if horizon > 0:
                # 小时预测：从 future_preds 列表取值
                future_seq = p.get('future_preds', [])
                for fi, fh_item in enumerate(future_seq[:total_pred_cols]):
                    fval = fh_item.get('predicted')
                    fconf = fh_item.get('confidence', 0)
                    cell_text = self._fmt_pred_val(fval)
                    if fval is not None and fconf > 0:
                        cell_text += f"\n({fconf*100:.0f}%)"
                    # 对单个时间点也做预警着色标记（文字颜色）
                    from PyQt5.QtGui import QBrush
                    from data_processor import get_warning_level
                    fi_level = get_warning_level(fval, p.get('param', ''), self.thresholds) if fval is not None else "正常"
                    fi_item = QTableWidgetItem(cell_text)
                    fi_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    # 超标单元格加深色背景（比整行高亮更细粒度）
                    fi_bg = self._warning_level_color(fi_level)
                    if fi_bg:
                        fi_item.setBackground(QColor(fi_bg))
                    self.pred_table.setItem(row_idx, 2 + fi, fi_item)
                # 如果 future_seq 长度不足（数据不够），剩余列填"—"
                for fi in range(len(future_seq), total_pred_cols):
                    item_dash = QTableWidgetItem("—")
                    item_dash.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.pred_table.setItem(row_idx, 2 + fi, item_dash)
            else:
                # 日均预测：只有1列
                cur_val = p.get('cur_pred') if 'cur_pred' in p else p.get('predicted')
                item_d = QTableWidgetItem(self._fmt_pred_val(cur_val))
                item_d.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.pred_table.setItem(row_idx, 2, item_d)

            # ── 趋势列 ───────────────────────────────────────────────────────
            trend      = p.get('trend', 'stable')
            trend_rate = p.get('trend_rate', 0)
            item_t = QTableWidgetItem(self._trend_text(trend, trend_rate))
            item_t.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.pred_table.setItem(row_idx, col_trend, item_t)

            # ── 置信度列 ─────────────────────────────────────────────────────
            confidence = p.get('confidence', 0)
            item_c = QTableWidgetItem(f"{confidence*100:.0f}%")
            item_c.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.pred_table.setItem(row_idx, col_conf, item_c)

            # ── 预警状态列 ───────────────────────────────────────────────────
            warn_level = p.get('warning_level', '正常')
            item_w = QTableWidgetItem(warn_text_map.get(warn_level, warn_level))
            item_w.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.pred_table.setItem(row_idx, col_warn, item_w)

            # ── 整行背景着色（预警行高亮） ───────────────────────────────────
            bg_hex = self._warning_level_color(warn_level)
            if bg_hex:
                bg = QColor(bg_hex)
                for col in [0, 1, col_trend, col_conf, col_warn]:
                    cell = self.pred_table.item(row_idx, col)
                    if cell:
                        cell.setBackground(bg)

        # ── 底部统计文字 ─────────────────────────────────────────────────────
        warn_count = sum(1 for p in predictions if p.get('is_warning'))
        total = len(predictions)
        if warn_count:
            summary = f"共 {total} 条预测，其中 {warn_count} 条存在超标风险 ⚠  |  声音报警：{'已开启 🔔' if self.sound_alarm_enabled else '已关闭 🔕'}"
        else:
            summary = f"共 {total} 条预测，未来 {horizon if horizon else 0} 小时内全部正常 ✅"
        self.pred_result_label.setText(summary)

    def _show_hour_prediction_table(self, predictions):
        """显示小时预测结果（动态多小时列）"""
        horizon = getattr(self, 'prediction_horizon', 3)
        self._fill_pred_table(predictions, horizon=horizon)
        if predictions:
            self._show_prediction_chart(predictions, "小时数据预测")
        else:
            self.pred_chart.clear()

    def _show_day_prediction_table(self, predictions):
        """显示日均预测结果（单列，无多小时）"""
        self._fill_pred_table(predictions, horizon=0)
        if predictions:
            self._show_prediction_chart(predictions, "当日均值预测")
        else:
            self.pred_chart.clear()

    def _get_supported_params_for_visible_subs(self):
        """收集当前可见排口（subs）真正支持的参数名称（去重，按 CODE_TO_NAME 的 key 顺序排列）。

        下拉框应展示当前企业所有排口支持的完整参数列表，而不只是已被预测过的参数。
        返回空列表表示无法获取（交由调用方兜底）。
        """
        if not getattr(self, 'current_grouped_data', None):
            return []

        supported = set()
        cache = getattr(self, '_sub_itemcodes_cache', {}) or {}
        for uk, sd in self.current_grouped_data.items():
            subid = sd.get('subid', uk)
            sub_type = sd.get('subtype', '')
            subtype_code = get_subtype_code(sub_type)
            subname = sd.get('subname', '')

            # 优先使用已缓存的监测项目代码，避免重复 API 调用
            cache_key = f"{subid}_{subtype_code}"
            codes = cache.get(cache_key) if isinstance(cache, dict) else None
            if not codes:
                try:
                    codes = self._get_sub_itemcodes(subid, subname, subtype_code)
                except Exception:
                    codes = ''
            if not codes:
                continue

            code_list = [c.strip() for c in codes.split(',') if c.strip()]
            for code in code_list:
                name = CODE_TO_NAME.get(code)
                if name:
                    supported.add(name)

        if not supported:
            return []

        # 按 CODE_TO_NAME 的 key 顺序输出（同名参数只取首次出现的 code，完成去重）
        ordered = []
        seen = set()
        for code in CODE_TO_NAME.keys():
            name = CODE_TO_NAME[code]
            if name in supported and name not in seen:
                seen.add(name)
                ordered.append(name)
        return ordered

    def _show_prediction_chart(self, predictions, title):
        """显示预测趋势图 —— 缓存全量预测数据，更新参数下拉框（展示全部支持参数），然后按选中参数绘图"""
        # 缓存全量预测结果和标题，供下拉框切换时复用
        self._pred_chart_all_predictions = predictions
        self._pred_chart_title = title

        # ── 更新下拉框：展示当前可见排口真正支持的全部参数（不论是否被预测）────
        supported_params = self._get_supported_params_for_visible_subs()
        if not supported_params:
            # 兜底：按原逻辑仅展示被预测过的参数，避免下拉框为空
            supported_params = []
            for p in predictions:
                pname = p.get('param', '')
                if pname and pname not in supported_params:
                    supported_params.append(pname)

        # 三项优化 T02：用可勾选"对比参数"面板替代原单参数下拉框
        self._populate_compare_list(supported_params)

        # 触发多指标对比绘制（保留单次预测点叠加逻辑）
        self._draw_pred_chart_multi(self._get_checked_compare_params(), self._pred_chart_mode)

    def _draw_pred_chart_for_param(self, param_name):
        """按指定参数，把所有支持该参数的排放口的历史曲线 + 预测值画在同一张图上。

        若该参数尚未被预测（例如用户在"选择预测指标"中只勾选了部分参数），
        则实时触发按需预测（不影响用户已选指标与预测表）：先展示历史曲线，
        预测完成后自动刷新并加入预测点。
        """
        predictions = getattr(self, '_pred_chart_all_predictions', []) or []
        title       = getattr(self, '_pred_chart_title', '预测趋势图')

        if not param_name:
            self.pred_chart.clear()
            return

        # 建立 (ent_name, subname) -> 预测记录 的索引，便于按排放口取预测值
        pred_by_sub = {}
        for p in predictions:
            if p.get('param') == param_name:
                key = (p.get('ent_name', ''), p.get('subname', ''))
                pred_by_sub[key] = p

        # 该参数没有任何预测结果，但存在可见排口 => 按需发起后台预测
        if not pred_by_sub and getattr(self, 'current_grouped_data', None):
            self._request_on_demand_prediction(param_name)

        from datetime import datetime, timedelta

        end_time   = datetime.now()
        start_time = end_time - timedelta(hours=24)
        start_str  = start_time.strftime("%Y-%m-%d %H:%M")
        end_str    = end_time.strftime("%Y-%m-%d %H:%M")

        # 遍历所有支持该参数的排放口，组成一条曲线
        series_list  = []
        common_times = None   # 用最长的时间轴作为公共 X 轴

        grouped = getattr(self, 'current_grouped_data', {}) or {}
        for uk, sd in grouped.items():
            subname  = sd.get('subname', '')
            ent_name = sd.get('ent_name', '')
            subid    = sd.get('subid', uk)
            sub_type = sd.get('subtype', '')
            subtype_code = get_subtype_code(sub_type)

            # 获取该排放口的监测项目代码
            try:
                codes = self._get_sub_itemcodes(subid, subname, subtype_code)
            except Exception:
                codes = None
            if not codes:
                continue

            # 找出 param_name 对应的 code（该排口不支持此参数则跳过）
            code_list    = [c.strip() for c in codes.split(',') if c.strip()]
            code_to_name = {code: CODE_TO_NAME.get(code, code) for code in code_list}
            target_code  = None
            for c, name in code_to_name.items():
                if name == param_name:
                    target_code = c
                    break
            if not target_code:
                continue

            # 获取 client
            client_info = self.multi_client.get_client_by_subid(subid)
            if not client_info:
                continue
            client = client_info["client"]

            try:
                result = client.query_history(
                    subid, subtype_code, codes,
                    start_str, end_str, index=1, page=1, rows=100
                )
                history_rows = result.get('rows', [])
            except Exception as e:
                print(f"[WARN] 获取 {subname} 历史数据失败: {e}")
                history_rows = []

            val_key = f"val_{target_code}"

            if history_rows:
                times_this  = [r.get('DateTime', '') for r in history_rows]
                values_hist = [float(r.get(val_key)) if r.get(val_key) else None
                               for r in history_rows]
            else:
                times_this  = []
                values_hist = []

            # 预测值（命中缓存则取，否则待按需预测完成后刷新）
            pred = pred_by_sub.get((ent_name, subname))
            pred_val = None
            if pred:
                pred_val = pred.get('predicted') if pred.get('predicted') is not None else pred.get('cur_pred')

            # 拼接时间轴和数据（历史 + "预测"标注点）
            times_with_pred  = times_this + ["预测"]
            values_with_pred = values_hist + [pred_val]

            if common_times is None or len(times_with_pred) > len(common_times):
                common_times = times_with_pred

            # 曲线名称：多企业时加企业名，单企业只显示排放口名
            curve_label = f"{ent_name}-{subname}" if ent_name else subname

            series_list.append({
                "name": curve_label,
                "data": values_with_pred,
                "times": times_with_pred,   # 保留各自的时间轴，ChartWidget 会用 name 匹配
            })

        if not series_list:
            # 无历史、无预测：清空图表
            self.pred_chart.clear()
            return

        # 如果各排放口时间轴长度不一致，对齐到最长时间轴（尾部补 None）
        max_len = max(len(s["times"]) for s in series_list)
        for s in series_list:
            diff = max_len - len(s["data"])
            if diff > 0:
                s["data"] = s["data"] + [None] * diff
        if common_times and len(common_times) < max_len:
            common_times = common_times + [""] * (max_len - len(common_times))

        chart_title = f"{title} — {param_name}（各排放口对比）"
        if not pred_by_sub:
            # 预测生成中，标题追加提示
            chart_title += "（预测生成中…）"
        try:
            self.pred_chart.plot_series(
                times=common_times or [],
                series_list=[{"name": s["name"], "data": s["data"]} for s in series_list],
                title=chart_title
            )
        except Exception as e:
            print(f"[ERROR] 绘制预测趋势图失败: {e}")
            import traceback
            traceback.print_exc()

    def _request_on_demand_prediction(self, param_name):
        """对下拉框中选中的、尚未预测的参数发起后台按需预测。

        仅刷新趋势图缓存，不会修改 self.prediction_params（用户已选指标），
        也不会重建预测表，从而满足"不选择则预测所有参数"的既有逻辑不被破坏。
        """
        if not getattr(self, 'current_grouped_data', None):
            return
        # 避免对同一参数重复发起后台任务
        if (getattr(self, '_on_demand_param', None) == param_name
                and getattr(self, '_pred_ondemand_worker', None) is not None):
            try:
                if self._pred_ondemand_worker.isRunning():
                    return
            except RuntimeError:
                self._pred_ondemand_worker = None

        pred_type = getattr(self, '_pred_chart_title', '小时数据预测') or '小时数据预测'
        self._on_demand_param = param_name

        worker = PredictionWorker(
            grouped_data=self.current_grouped_data,
            multi_client=self.multi_client,
            prediction_horizon=self.prediction_horizon,
            prediction_params=[param_name],
            pred_type=pred_type,
            thresholds=self.thresholds,
            sub_itemcodes_cache=self._sub_itemcodes_cache,
            api_data_cache=self._api_data_cache,
            thresholds_cache=getattr(self, '_thresholds_cache', {}),
            get_subtype_code_fn=get_subtype_code,
            is_water_sub_fn=is_water_sub,
            code_to_name_map=CODE_TO_NAME,
            intervention_sm=getattr(self, '_intervention_sm', None),
            get_sub_itemcodes_fn=self._get_sub_itemcodes,
        )
        worker.prediction_done.connect(
            lambda preds: self._on_ondemand_prediction_done(preds, param_name)
        )
        worker.finished.connect(worker.deleteLater)
        self._pred_ondemand_worker = worker
        worker.start()
        print(f"[PRED] 按需预测已发起: {param_name}")

    def _on_ondemand_prediction_done(self, new_preds, param_name):
        """按需预测完成：合并进趋势图缓存并重绘（仅刷新趋势图，不影响预测表）。"""
        try:
            existing = getattr(self, '_pred_chart_all_predictions', []) or []
            # 用新结果替换该参数的旧预测，避免重复
            merged = [p for p in existing if p.get('param') != param_name]
            merged.extend(new_preds)
            self._pred_chart_all_predictions = merged

            # 三项优化 T02：若该参数仍在勾选列表内，则刷新多指标对比图
            if param_name in self._get_checked_compare_params():
                self._draw_pred_chart_multi(self._get_checked_compare_params(), self._pred_chart_mode)
        except Exception as e:
            print(f"[ERROR] 按需预测结果处理失败: {e}")
            import traceback
            traceback.print_exc()

    # ══════════════════════════════════════════════════════════════════════════
    # 三项优化 T02：多指标对比 UI + 后台取数（替代原单参数下拉框）
    # ══════════════════════════════════════════════════════════════════════════

    def _populate_compare_list(self, params):
        """用可见排口支持的参数填充可勾选对比面板。

        保留已有勾选；面板首次为空时默认全勾选（给出初始图表）。
        """
        lst = getattr(self, 'pred_param_list', None)
        if lst is None:
            return
        prev_checked = set()
        for i in range(lst.count()):
            it = lst.item(i)
            if it.checkState() == Qt.Checked:
                prev_checked.add(it.text())
        auto_check_all = (lst.count() == 0)
        lst.blockSignals(True)
        lst.clear()
        for p in params:
            it = QListWidgetItem(p)
            it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
            if prev_checked:
                it.setCheckState(Qt.Checked if p in prev_checked else Qt.Unchecked)
            else:
                it.setCheckState(Qt.Checked if auto_check_all else Qt.Unchecked)
            lst.addItem(it)
        lst.blockSignals(False)

    def _get_checked_compare_params(self):
        """返回当前勾选的对比参数名列表。"""
        lst = getattr(self, 'pred_param_list', None)
        if lst is None:
            return []
        return [lst.item(i).text() for i in range(lst.count())
                if lst.item(i).checkState() == Qt.Checked]

    def _on_compare_params_changed(self):
        """勾选变化：重算并触发取数/重绘。"""
        checked = self._get_checked_compare_params()
        self._draw_pred_chart_multi(checked, self._pred_chart_mode)

    def _on_pred_mode_changed(self, mode):
        """模式切换：记录并重绘。"""
        self._pred_chart_mode = mode
        checked = self._get_checked_compare_params()
        self._draw_pred_chart_multi(checked, mode)

    def _on_normalize_toggled(self, flag):
        """归一化开关：记录并重绘。"""
        self._pred_chart_normalize = bool(flag)
        try:
            self.pred_normalize_btn.setText(f"归一化: {'开' if flag else '关'}")
        except Exception:
            pass
        checked = self._get_checked_compare_params()
        self._draw_pred_chart_multi(checked, self._pred_chart_mode)

    def _select_all_params(self):
        lst = getattr(self, 'pred_param_list', None)
        if lst is None:
            return
        lst.blockSignals(True)
        for i in range(lst.count()):
            lst.item(i).setCheckState(Qt.Checked)
        lst.blockSignals(False)
        self._on_compare_params_changed()

    def _invert_params(self):
        lst = getattr(self, 'pred_param_list', None)
        if lst is None:
            return
        lst.blockSignals(True)
        for i in range(lst.count()):
            it = lst.item(i)
            it.setCheckState(Qt.Unchecked if it.checkState() == Qt.Checked else Qt.Checked)
        lst.blockSignals(False)
        self._on_compare_params_changed()

    def _clear_params(self):
        lst = getattr(self, 'pred_param_list', None)
        if lst is None:
            return
        lst.blockSignals(True)
        for i in range(lst.count()):
            lst.item(i).setCheckState(Qt.Unchecked)
        lst.blockSignals(False)
        self._on_compare_params_changed()

    def _build_hist_tasks(self, checked_params, mode):
        """按 docs/system_design.md §5 裁定生成 sub_tasks（按 sub 聚合）。

        每个 sub 仅发起 1 次历史请求（取全量 codes），本地按 params 拆分多参数。
        Returns: list[dict]，元素含 subid/subtype_code/codes/params/start/end/
                 index/use_corrected/subname/ent_name。
        """
        grouped = getattr(self, 'current_grouped_data', {}) or {}
        if not grouped or not checked_params:
            return []
        checked_set = set(checked_params)

        # ── 确定参与对比的排口集合 ──────────────────────────────────────────
        if mode == "同排口多参数":
            target = None
            sel = getattr(self, 'selected_subid', None)
            if sel is not None:
                for uk, sd in grouped.items():
                    if sd.get('subid', uk) == sel:
                        target = (uk, sd)
                        break
            if target is None:
                # 无选中排口：取首个可见排口
                for uk, sd in grouped.items():
                    target = (uk, sd)
                    break
            subs = [target] if target else []
        else:  # 同参数多排口：所有可见排口
            subs = list(grouped.items())

        sub_tasks = []
        for uk, sd in subs:
            subid = sd.get('subid', uk)
            sub_type = sd.get('subtype', '')
            subtype_code = get_subtype_code(sub_type)
            subname = sd.get('subname', '')
            ent_name = sd.get('ent_name', '')

            try:
                codes = self._get_sub_itemcodes(subid, subname, subtype_code)
            except Exception:
                codes = None
            if not codes:
                continue
            code_list = [c.strip() for c in codes.split(',') if c.strip()]
            code_to_name = {code: CODE_TO_NAME.get(code, code) for code in code_list}

            # 该排口支持且被勾选的参数 -> (code, name, axis)
            params = []
            for code, name in code_to_name.items():
                if name in checked_set:
                    params.append((code, name, classify_axis(name)))
            if not params:
                continue

            # 时间窗：近 24h（与设计 §8 一致）
            end_time = datetime.now()
            start_time = end_time - timedelta(hours=24)
            start_str = start_time.strftime("%Y-%m-%d %H:%M")
            end_str = end_time.strftime("%Y-%m-%d %H:%M")
            use_corrected = '热电' in ent_name

            sub_tasks.append({
                'subid': subid,
                'subtype_code': subtype_code,
                'codes': ','.join(code_list),
                'params': params,
                'start': start_str,
                'end': end_str,
                'index': HISTORY_INDEX,
                'use_corrected': use_corrected,
                'subname': subname,
                'ent_name': ent_name,
            })
        return sub_tasks

    def _cached_hist_payload(self, task):
        """若历史缓存命中且未过期，返回与 worker 一致的 history_result payload；否则 None。"""
        key = make_history_cache_key(
            task['subid'], task['subtype_code'], task['codes'],
            task['start'], task['end'], task['index'], task['use_corrected'])
        entry = self._history_cache.get(key)
        if entry is None:
            return None
        ts, rows = entry
        if time.time() - ts >= HISTORY_TTL:
            return None
        if not isinstance(rows, list):
            return None
        series = []
        for (code, name, axis) in task['params']:
            val_key = f"val_{code}"
            times = [r.get('DateTime', '') for r in rows]
            values = []
            for r in rows:
                v = r.get(val_key)
                if v not in (None, ''):
                    try:
                        values.append(float(v))
                    except (ValueError, TypeError):
                        values.append(None)
                else:
                    values.append(None)
            series.append({"param_name": name, "code": code,
                           "axis": axis, "times": times, "values": values})
        return {
            "subid": task['subid'], "subname": task['subname'],
            "ent_name": task['ent_name'], "subtype_code": task['subtype_code'],
            "window": f"{task['start']} ~ {task['end']}", "series": series,
        }

    def _draw_pred_chart_multi(self, checked_params, mode):
        """多指标对比主绘制入口（T02 核心）。

        流程：
          1. 生成 sub_tasks（_build_hist_tasks）
          2. 预检 _history_cache，命中者即时成 series
          3. 未命中者交给后台 HistoryFetchWorker 取数（取消上一个未完成 worker）
          4. history_result 累加 / fetch_finished 时组装 left+right series、归一化、
             追加预测点 -> pred_chart.plot_series -> canvas.draw_idle()
        全程后台取数，主线程零阻塞（消除切指标卡顿）。
        """
        grouped = getattr(self, 'current_grouped_data', {}) or {}
        if not checked_params or not grouped:
            try:
                self.pred_chart.clear()
            except Exception:
                pass
            return

        # 取消上一个未完成 worker，避免污染 _hist_results
        if self._hist_worker is not None and self._hist_worker.isRunning():
            try:
                self._hist_worker.quit()
                self._hist_worker.wait()
            except Exception:
                pass
            self._hist_worker = None

        sub_tasks = self._build_hist_tasks(checked_params, mode)
        if not sub_tasks:
            try:
                self.pred_chart.clear()
            except Exception:
                pass
            return

        # 预检缓存，命中即时成图，未命中送 worker
        self._hist_results = {}
        uncached = []
        for task in sub_tasks:
            payload = self._cached_hist_payload(task)
            if payload is not None:
                self._hist_results[task['subid']] = payload
            else:
                uncached.append(task)

        if not uncached:
            # 全部命中缓存：直接组装绘制
            self._assemble_and_plot_pred_chart(checked_params, mode)
            return

        # 启动后台取数
        self._pred_chart_loading = True
        worker = HistoryFetchWorker(self.multi_client, uncached,
                                    self._history_cache, HISTORY_TTL)
        wref = worker
        worker.history_result.connect(self._on_hist_result)
        worker.fetch_finished.connect(
            lambda: self._on_hist_fetch_finished(checked_params, mode, wref))
        worker.finished.connect(worker.deleteLater)
        self._hist_worker = worker
        worker.start()

    def _on_hist_result(self, payload):
        """后台逐 sub 回传：累加历史 series。"""
        subid = payload.get('subid')
        if subid is not None:
            self._hist_results[subid] = payload

    def _on_hist_fetch_finished(self, checked_params, mode, worker):
        """后台取数完成：组装并绘制。"""
        if self._hist_worker is not worker:
            # 已有新的取数任务，放弃本次过期结果
            return
        self._hist_worker = None
        self._pred_chart_loading = False
        self._assemble_and_plot_pred_chart(checked_params, mode)

    def _assemble_and_plot_pred_chart(self, checked_params, mode):
        """从 _hist_results 组装 left/right series、追加预测点、归一化、截断并绘制。"""
        results = getattr(self, '_hist_results', {}) or {}
        if not results:
            try:
                self.pred_chart.clear()
            except Exception:
                pass
            return

        # 预测点索引：(ent_name, subname, param) -> 预测值
        predictions = getattr(self, '_pred_chart_all_predictions', []) or []
        pred_by_sub_param = {}
        for p in predictions:
            key = (p.get('ent_name', ''), p.get('subname', ''), p.get('param', ''))
            val = p.get('predicted')
            if val is None:
                val = p.get('cur_pred')
            pred_by_sub_param[key] = val

        # 多企业消歧
        grouped = getattr(self, 'current_grouped_data', {}) or {}
        ent_names = set(sd.get('ent_name', '') for sd in grouped.values())
        multi_ent = len(ent_names) > 1

        # 组装原始曲线
        raw_curves = []  # {name, axis, data, times}
        for subid, payload in results.items():
            subname = payload.get('subname', '')
            ent_name = payload.get('ent_name', '')
            for s in payload.get('series', []):
                param = s.get('param_name', '')
                pred_val = pred_by_sub_param.get((ent_name, subname, param))
                times_cur = list(s.get('times', [])) + ["预测"]
                data_cur = list(s.get('values', [])) + [pred_val]
                label = f"{param}@{ent_name}-{subname}" if multi_ent else f"{param}@{subname}"
                raw_curves.append({
                    "name": label,
                    "axis": s.get('axis', 'left'),
                    "data": data_cur,
                    "times": times_cur,
                })

        # 触发缺失参数的按需预测（保留单次预测点叠加逻辑）
        requested = set()
        for subid, payload in results.items():
            subname = payload.get('subname', '')
            ent_name = payload.get('ent_name', '')
            for s in payload.get('series', []):
                param = s.get('param_name', '')
                if (ent_name, subname, param) not in pred_by_sub_param and param not in requested:
                    requested.add(param)
                    self._request_on_demand_prediction(param)

        if not raw_curves:
            try:
                self.pred_chart.clear()
            except Exception:
                pass
            return

        # SERIES_CAP 截断
        hidden_count = 0
        if len(raw_curves) > SERIES_CAP:
            hidden_count = len(raw_curves) - SERIES_CAP
            raw_curves = raw_curves[:SERIES_CAP]

        # 对齐到公共 X 轴（最长曲线）
        max_len = max(len(c['data']) for c in raw_curves)
        common_times = []
        for c in raw_curves:
            if len(c['data']) == max_len:
                common_times = list(c['times'])
                break
        if not common_times:
            common_times = ["" for _ in range(max_len)]
        for c in raw_curves:
            diff = max_len - len(c['data'])
            if diff > 0:
                c['data'] = c['data'] + [None] * diff
                c['times'] = c['times'] + [""] * diff

        normalize = bool(self._pred_chart_normalize)
        if normalize:
            left_series = [{"name": c['name'], "data": c['data']} for c in raw_curves]
            right_series = None
            right_ylabel = ""
        else:
            left_series = [{"name": c['name'], "data": c['data']}
                           for c in raw_curves if c['axis'] == 'left']
            right_series = [{"name": c['name'], "data": c['data']}
                            for c in raw_curves if c['axis'] == 'right']
            right_series = right_series if right_series else None
            right_ylabel = "pH / 温度 / 水温"

        title = f"{getattr(self, '_pred_chart_title', '预测趋势图')} — {mode}"
        if hidden_count > 0:
            title += f"（已隐藏 {hidden_count} 条）"

        try:
            self.pred_chart.plot_series(
                times=common_times,
                series_list=left_series,
                title=title,
                right_series_list=right_series,
                right_ylabel=right_ylabel,
                normalize=normalize,
            )
            self.pred_chart.canvas.draw_idle()
        except Exception as e:
            print(f"[ERROR] 绘制多指标预测趋势图失败: {e}")
            import traceback
            traceback.print_exc()

    def _load_thresholds(self):
        """加载预警阈值（支持上下限）"""
        self.threshold_table.setRowCount(len(self.thresholds))
        for row, (param_name, threshold) in enumerate(self.thresholds.items()):
            self.threshold_table.setItem(row, 0, QTableWidgetItem(param_name))
            
            # 解析阈值格式：支持单个值(上限)或元组(下限,上限)
            if isinstance(threshold, tuple):
                lower, upper = threshold
            else:
                lower, upper = None, threshold
            
            self.threshold_table.setItem(row, 1, QTableWidgetItem(str(lower) if lower is not None else "无"))
            self.threshold_table.setItem(row, 2, QTableWidgetItem(str(upper) if upper is not None else "无"))

            edit_btn = QPushButton("修改")
            edit_btn.clicked.connect(lambda _, r=row: self._edit_threshold(r))
            self.threshold_table.setCellWidget(row, 3, edit_btn)

    def _edit_threshold(self, row):
        """编辑预警阈值（支持上下限，修改后立即生效）"""
        param_name = self.threshold_table.item(row, 0).text()
        
        # 获取当前阈值
        threshold = self.thresholds.get(param_name)
        if isinstance(threshold, tuple):
            current_lower, current_upper = threshold
        else:
            current_lower, current_upper = None, threshold

        dialog = QDialog(self)
        dialog.setWindowTitle(f"修改预警阈值 - {param_name}")
        dialog.setFixedSize(350, 220)
        dialog.setStyleSheet(f"""
            QDialog {{
                background: {COLORS['bg_dark']};
                color: {COLORS['text_primary']};
            }}
            QLabel {{
                color: {COLORS['text_primary']};
                padding: 5px;
            }}
            QLineEdit {{
                background: {COLORS['bg_input']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 4px;
                padding: 8px;
            }}
        """)
        layout = QFormLayout()
        
        # 下限输入
        lower_edit = QLineEdit()
        lower_edit.setPlaceholderText("无下限请留空")
        if current_lower is not None:
            lower_edit.setText(str(current_lower))
        layout.addRow("下限值:", lower_edit)
        
        # 上限输入
        upper_edit = QLineEdit()
        upper_edit.setPlaceholderText("无上限请留空")
        if current_upper is not None:
            upper_edit.setText(str(current_upper))
        layout.addRow("上限值:", upper_edit)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        dialog.setLayout(layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            # 解析新值
            new_lower = None
            new_upper = None
            try:
                lower_text = lower_edit.text().strip()
                if lower_text:
                    new_lower = float(lower_text)
            except ValueError:
                pass
            
            try:
                upper_text = upper_edit.text().strip()
                if upper_text:
                    new_upper = float(upper_text)
            except ValueError:
                pass
            
            # 验证：至少要有一个值
            if new_lower is None and new_upper is None:
                QMessageBox.warning(self, "错误", "下限和上限不能同时为空")
                return
            
            # 保存格式：如果只有上限，存单个值；否则存元组
            if new_lower is None:
                self.thresholds[param_name] = new_upper
            elif new_upper is None:
                self.thresholds[param_name] = new_lower
            else:
                self.thresholds[param_name] = (new_lower, new_upper)
            
            # 更新表格显示
            self.threshold_table.setItem(row, 1, QTableWidgetItem(str(new_lower) if new_lower is not None else "无"))
            self.threshold_table.setItem(row, 2, QTableWidgetItem(str(new_upper) if new_upper is not None else "无"))
            # 保存并立即生效
            if save_warn_thresholds(self.thresholds):
                print(f"[INFO] 预警阈值已更新: {param_name} = {edit.value()}")
                self._apply_thresholds_now()
            else:
                QMessageBox.warning(self, "错误", "预警阈值保存失败，请重试")

    def _apply_thresholds_now(self):
        """使用当前 self.thresholds 立即重新检查预警并刷新界面状态"""
        if not (hasattr(self, 'current_grouped_data') and self.current_grouped_data):
            print("[INFO] 暂无实时数据缓存，新阈值将在下次刷新时生效")
            return
        try:
            # 重新运行预警检查
            self.warning_system.check_and_alert(self.current_grouped_data, self.thresholds)
            # 刷新排放口列表颜色/状态
            self._update_sub_list(self.current_grouped_data)
            # 如果当前有选中排口，也刷新右侧实时数据面板
            if hasattr(self, 'selected_unique_key'):
                self._show_sub_realtime(self.selected_unique_key)
            # 更新今日预警统计
            if hasattr(self, '_today_warning_label'):
                today_count = self.warning_system.get_warning_count_today()
                self._today_warning_label.setText(f"📊 今日预警次数: {today_count}")
            print("[INFO] 新阈值已立即生效，界面已刷新")
        except Exception as e:
            print(f"[WARN] 立即应用阈值时出错: {e}")

    def _save_thresholds(self):
        """保存预警阈值 —— 从表格读回最新值（支持上下限），保存并立即生效"""
        # 1. 从表格读取所有行的最新值（用户可能直接在表格里编辑了单元格）
        for row in range(self.threshold_table.rowCount()):
            name_item = self.threshold_table.item(row, 0)
            lower_item = self.threshold_table.item(row, 1)
            upper_item = self.threshold_table.item(row, 2)
            
            if name_item:
                param_name = name_item.text().strip()
                
                # 解析下限
                new_lower = None
                if lower_item:
                    lower_text = lower_item.text().strip()
                    if lower_text and lower_text != "无":
                        try:
                            new_lower = float(lower_text)
                        except ValueError:
                            pass
                
                # 解析上限
                new_upper = None
                if upper_item:
                    upper_text = upper_item.text().strip()
                    if upper_text and upper_text != "无":
                        try:
                            new_upper = float(upper_text)
                        except ValueError:
                            pass
                
                # 保存格式：如果只有下限，存单个值（只有下限的少数参数如烟气含氧量）
                # 如果只有上限，存单个值
                # 如果上下限都有，存元组
                if new_lower is not None and new_upper is not None:
                    self.thresholds[param_name] = (new_lower, new_upper)
                elif new_lower is not None:
                    self.thresholds[param_name] = new_lower
                elif new_upper is not None:
                    self.thresholds[param_name] = new_upper

        # 2. 持久化到文件
        if save_warn_thresholds(self.thresholds):
            # 3. 立即用新阈值重新检查并刷新界面
            self._apply_thresholds_now()
            QMessageBox.information(self, "成功", "预警阈值已保存并立即生效！")
        else:
            QMessageBox.warning(self, "错误", "保存失败，请重试")

    def _setup_menu(self):
        """设置菜单栏"""
        menubar = self.menuBar()
        menubar.setStyleSheet(f"""
            QMenuBar {{
                background: {COLORS['bg_dark']};
                color: {COLORS['text_primary']};
                border-bottom: 2px solid {COLORS['primary']};
                padding: 5px;
            }}
            QMenuBar::item {{
                background: transparent;
                padding: 8px 15px;
                border-radius: 4px;
            }}
            QMenuBar::item:selected {{
                background: {COLORS['primary']};
                color: white;
            }}
            QMenu {{
                background: {COLORS['bg_card']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
            }}
            QMenu::item {{
                padding: 8px 30px;
                border-radius: 4px;
            }}
            QMenu::item:selected {{
                background: {COLORS['primary']};
                color: white;
            }}
        """)

        # 文件菜单
        file_menu = menubar.addMenu("📁 文件")

        # 重新登录
        relogin_action = file_menu.addAction("🔄 重新登录")
        relogin_action.triggered.connect(self._relogin)

        file_menu.addSeparator()

        # 退出
        exit_action = file_menu.addAction("🚪 退出")
        exit_action.triggered.connect(self.close)

        # 设置菜单
        settings_menu = menubar.addMenu("⚙️ 设置")

        # 修改密码
        change_password_action = settings_menu.addAction("🔑 修改密码")
        change_password_action.triggered.connect(self._change_password)

        # 帮助菜单
        help_menu = menubar.addMenu("❓ 帮助")

        # 关于
        about_action = help_menu.addAction("ℹ️ 关于")
        about_action.triggered.connect(self._show_about)

    def _relogin(self):
        """重新登录"""
        reply = QMessageBox.question(
            self,
            "重新登录",
            "确定要重新登录吗？\n当前窗口将关闭并重新显示登录界面。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            # 关闭当前窗口，程序会自动显示登录界面
            self.close()
            # 重新启动登录对话框
            from main import _login_dialog, _main_window
            global _login_dialog, _main_window
            _main_window = None  # 清除引用
            _login_dialog.show()

    def _change_password(self):
        """修改密码"""
        from account_manager import load_accounts, save_accounts

        accounts = load_accounts()

        if not accounts:
            QMessageBox.information(self, "提示", "当前没有可修改的账户")
            return

        # 创建账户选择对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("修改密码")
        dialog.setFixedSize(450, 300)
        dialog.setStyleSheet(f"""
            QDialog {{
                background: {COLORS['bg_dark']};
                color: {COLORS['text_primary']};
            }}
            QGroupBox {{
                border: 2px solid {COLORS['border']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                color: {COLORS['accent']};
            }}
            QLabel {{
                color: {COLORS['text_primary']};
                padding: 5px;
            }}
            QLineEdit, QComboBox {{
                background: {COLORS['bg_input']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 4px;
                padding: 8px;
            }}
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['primary']}, stop:1 {COLORS['secondary']});
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['secondary']}, stop:1 {COLORS['primary']});
            }}
        """)

        layout = QVBoxLayout()
        form_layout = QFormLayout()

        # 账户选择
        account_combo = QComboBox()
        for acc in accounts:
            account_combo.addItem(acc['name'])
        form_layout.addRow("选择账户:", account_combo)

        # 新密码输入
        new_password = QLineEdit()
        new_password.setEchoMode(QLineEdit.EchoMode.Password)
        form_layout.addRow("新密码:", new_password)

        # 确认密码
        confirm_password = QLineEdit()
        confirm_password.setEchoMode(QLineEdit.EchoMode.Password)
        form_layout.addRow("确认密码:", confirm_password)

        layout.addLayout(form_layout)

        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        ok_btn = QPushButton("✓ 确认修改")
        ok_btn.setMinimumWidth(100)
        ok_btn.clicked.connect(dialog.accept)
        button_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("✗ 取消")
        cancel_btn.setMinimumWidth(100)
        cancel_btn.clicked.connect(dialog.reject)
        button_layout.addWidget(cancel_btn)

        layout.addLayout(button_layout)
        dialog.setLayout(layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            account_idx = account_combo.currentIndex()
            new_pwd = new_password.text()
            confirm_pwd = confirm_password.text()

            if not new_pwd:
                QMessageBox.warning(self, "错误", "新密码不能为空")
                return

            if new_pwd != confirm_pwd:
                QMessageBox.warning(self, "错误", "两次输入的密码不一致")
                return

            if len(new_pwd) < 6:
                QMessageBox.warning(self, "错误", "密码长度不能少于6位")
                return

            # 更新密码
            accounts[account_idx]['password'] = new_pwd
            if save_accounts(accounts):
                QMessageBox.information(self, "成功", "密码修改成功！")
            else:
                QMessageBox.warning(self, "错误", "密码修改失败")

    def _show_about(self):
        """显示关于信息"""
        QMessageBox.information(
            self,
            "关于 GZ安环监测系统",
            f"版本：v1.0.0\n\n"
            f"GZ安环监测系统 - 跨电脑部署版本\n\n"
            f"功能特点：\n"
            f"• 五个企业账户一键登录\n"
            f"• 实时数据监测\n"
            f"• 历史数据查询\n"
            f"• 数据预测分析\n"
            f"• 预警阈值设置\n"
            f"• 完全便携，跨电脑使用\n\n"
            f"开发日期：2026-03-21"
        )

    def closeEvent(self, event):
        """关闭事件"""
        self.warning_system.stop_alert()
        event.accept()

    def _show_warning_history(self):
        """显示预警历史对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("预警历史记录")
        dialog.setFixedSize(900, 600)
        dialog.setStyleSheet(f"""
            QDialog {{
                background: {COLORS['bg_dark']};
                color: {COLORS['text_primary']};
            }}
            QGroupBox {{
                border: 2px solid {COLORS['border']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                color: {COLORS['secondary']};
            }}
            QTableWidget {{
                background: {COLORS['bg_input']};
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
            }}
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['primary']}, stop:1 {COLORS['secondary']});
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {COLORS['secondary']}, stop:1 {COLORS['primary']});
            }}
        """)

        layout = QVBoxLayout()

        # 工具栏
        toolbar = QHBoxLayout()
        toolbar.addWidget(QLabel("📋 预警历史记录（最近100条）"))
        toolbar.addStretch()

        clear_btn = QPushButton("🗑️ 清空历史")
        clear_btn.clicked.connect(lambda: self._clear_warning_history(dialog))
        toolbar.addWidget(clear_btn)

        export_btn = QPushButton("📥 导出")
        export_btn.clicked.connect(lambda: self._export_warning_history())
        toolbar.addWidget(export_btn)

        layout.addLayout(toolbar)

        # 预警历史表格
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels([
            "时间", "企业", "排放口", "参数", "数值", "预警等级"
        ])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)

        history = self.warning_system.get_warning_history(100)
        table.setRowCount(len(history))

        for row, w in enumerate(history):
            table.setItem(row, 0, QTableWidgetItem(w.get('datetime', '')))
            table.setItem(row, 1, QTableWidgetItem(w.get('ent_name', '')))
            table.setItem(row, 2, QTableWidgetItem(w.get('subname', '')))
            table.setItem(row, 3, QTableWidgetItem(w.get('param', '')))

            value_item = QTableWidgetItem(str(w.get('value', '')))
            table.setItem(row, 4, value_item)

            level_item = QTableWidgetItem(w.get('level', ''))
            # 根据预警等级设置颜色
            if w.get('level') == "红色预警":
                level_item.setForeground(QColor("#dc2626"))
            elif w.get('level') == "橙色预警":
                level_item.setForeground(QColor("#ea580c"))
            elif w.get('level') == "黄色预警":
                level_item.setForeground(QColor("#ca8a04"))
            table.setItem(row, 5, level_item)

        layout.addWidget(table)

        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.accept)
        close_btn.setMinimumWidth(100)
        close_btn.setMaximumHeight(40)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        dialog.setLayout(layout)
        dialog.exec()

    def _clear_warning_history(self, dialog):
        """清空预警历史"""
        reply = QMessageBox.question(
            dialog, "确认", "确定要清空所有预警历史记录吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.warning_system.clear_history()
            dialog.accept()
            QMessageBox.information(self, "成功", "预警历史已清空")

    def _export_warning_history(self):
        """导出预警历史"""
        history = self.warning_system.get_warning_history()
        if not history:
            QMessageBox.information(self, "提示", "暂无预警历史可导出")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出预警历史", "预警历史.xlsx", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            import pandas as pd
            import openpyxl

            df = pd.DataFrame(history)
            df.to_excel(file_path, index=False, engine='openpyxl')
            QMessageBox.information(self, "成功", f"已导出 {len(history)} 条预警记录到:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出预警历史时发生错误:\n{str(e)}")



